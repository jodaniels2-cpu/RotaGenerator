import io
import re
from datetime import datetime, date, time, timedelta

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================================================
# Rota Generator — Clean Rebuild v10d
# =========================================================
# Goals (per your rules):
# - Open Mon–Fri 08:00–18:30, 30-min slots
# - Front Desk: SLGP/JEN/BGS each must have exactly 1 person at a time
# - Nobody on Front Desk > 2.5h at a time (blocks 2h–3h, max 3h satisfies)
# - Triage Admin: SLGP and JEN only, covered until 16:00; max 3h blocks
#   **STRICT home-site only** (no cross-site triage)
# - Email box manned all day by JEN/BGS only (SLGP only if absolutely forced by rule? -> here: no)
# - Phones: minimum 2 people all day; Mon/Fri prioritise Phones over EMIS/Docman
# - Bookings: minimum 3 people all day; SLGP should be mostly bookings (enforced via filler bias)
# - Awaiting Response/PSA Admin: mandatory 10:00–16:00; optional 16:00–18:30
#   Site: Mon/Fri SLGP, Tue/Thu JEN, Wed BGS
# - Breaks: if shift > 6h, 30-min break scheduled as close to midpoint as possible, start times around 12:00–14:00
# - Minimum task block length 2h (except end-of-shift remainder can be <2h)
# - Weights (0–5) optional on Staff sheet, used as TIE-BREAK ONLY
# - Outputs: Weekly tabs with Grid, CoverageGaps, Totals, Timelines_All, and Timelines by site

# ---------------- Password (optional) ----------------
def require_password():
    pw = st.secrets.get("APP_PASSWORD")
    if not pw:
        return True
    if st.session_state.get("authed"):
        return True
    with st.form("login"):
        entered = st.text_input("Password", type="password")
        ok = st.form_submit_button("Log in")
        if ok and entered == pw:
            st.session_state.authed = True
            st.success("Logged in.")
            return True
        if ok:
            st.error("Incorrect password.")
    st.stop()

# ---------------- Parsing helpers ----------------
def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s).strip().lower())

def find_sheet(xls: pd.ExcelFile, candidates):
    names = {normalize(n): n for n in xls.sheet_names}
    for c in candidates:
        k = normalize(c)
        if k in names:
            return names[k]
    for n in xls.sheet_names:
        nn = normalize(n)
        for c in candidates:
            if normalize(c) in nn:
                return n
    return None

def pick_col(df: pd.DataFrame, candidates, required=True):
    cols = {normalize(c): c for c in df.columns}
    for cand in candidates:
        k = normalize(cand)
        if k in cols:
            return cols[k]
    for c in df.columns:
        nc = normalize(c)
        for cand in candidates:
            if normalize(cand) in nc:
                return c
    if required:
        raise KeyError(f"Missing required column among {candidates}. Available: {list(df.columns)}")
    return None

def to_time(x):
    if pd.isna(x) or x == "":
        return None
    if isinstance(x, time):
        return x
    if isinstance(x, datetime):
        return x.time()
    if isinstance(x, (float, int)):
        seconds = int(round(float(x) * 86400))
        return (datetime(2000, 1, 1) + timedelta(seconds=seconds)).time()
    return pd.to_datetime(str(x)).time()

def to_date(x):
    if pd.isna(x) or x == "":
        return None
    if isinstance(x, (datetime, pd.Timestamp)):
        return x.date()
    if isinstance(x, date):
        return x
    return pd.to_datetime(x).date()

def dt_of(d: date, t: time) -> datetime:
    return datetime(d.year, d.month, d.day, t.hour, t.minute)

def add_minutes(t: time, mins: int) -> time:
    return (datetime(2000, 1, 1, t.hour, t.minute) + timedelta(minutes=mins)).time()

def t_in_range(t: time, a: time, b: time) -> bool:
    return (t >= a) and (t < b)

def ensure_monday(d: date) -> date:
    return d - timedelta(days=d.weekday())

# ---------------- Read template (NO Parameters required) ----------------
def read_template(uploaded_bytes: bytes):
    xls = pd.ExcelFile(io.BytesIO(uploaded_bytes))
    staff_sheet = find_sheet(xls, ["Staff", "Skills", "People"])
    hours_sheet = find_sheet(xls, ["WorkingHours", "Hours", "Availability"])
    hols_sheet = find_sheet(xls, ["Holidays", "Leave", "Absence"])

    if not staff_sheet:
        raise ValueError(f"Could not find Staff sheet. Found: {xls.sheet_names}")
    if not hours_sheet:
        raise ValueError(f"Could not find WorkingHours sheet. Found: {xls.sheet_names}")

    staff_df = pd.read_excel(xls, sheet_name=staff_sheet)
    hours_df = pd.read_excel(xls, sheet_name=hours_sheet)
    hols_df = pd.read_excel(xls, sheet_name=hols_sheet) if hols_sheet else pd.DataFrame()

    name_c = pick_col(staff_df, ["Name", "StaffName"])
    home_c = pick_col(staff_df, ["HomeSite", "Site", "BaseSite"], required=False)

    staff_df = staff_df.copy()
    staff_df["Name"] = staff_df[name_c].astype(str).str.strip()
    staff_df["HomeSite"] = staff_df[home_c].astype(str).str.strip().str.upper() if home_c else ""

    def yn(v):
        if pd.isna(v):
            return False
        s = str(v).strip().lower()
        return s in ["y", "yes", "true", "1"]

    for c in staff_df.columns:
        if normalize(c).startswith("can") or normalize(c) in ["iscarolchurchyn", "iscarolchurch"]:
            staff_df[c] = staff_df[c].apply(yn)

    carol_c = pick_col(staff_df, ["IsCarolChurch(Y/N)", "IsCarolChurch", "Carol"], required=False)
    if carol_c:
        staff_df["IsCarolChurch"] = staff_df[carol_c].apply(bool)
    else:
        staff_df["IsCarolChurch"] = staff_df["Name"].str.lower().eq("carol church")

    hours_df = hours_df.copy()
    hours_name_c = pick_col(hours_df, ["Name", "StaffName"])
    hours_df["Name"] = hours_df[hours_name_c].astype(str).str.strip()

    for d in ["Mon", "Tue", "Wed", "Thu", "Fri"]:
        sc = pick_col(hours_df, [f"{d}Start", f"{d} Start", f"{d}_Start"], required=False)
        ec = pick_col(hours_df, [f"{d}End", f"{d} End", f"{d}_End"], required=False)
        hours_df[f"{d}Start"] = hours_df[sc].apply(to_time) if sc else None
        hours_df[f"{d}End"] = hours_df[ec].apply(to_time) if ec else None

    hols = []
    if not hols_df.empty:
        hn = pick_col(hols_df, ["Name", "StaffName"], required=False) or hols_df.columns[0]
        hs = pick_col(hols_df, ["StartDate", "Start"], required=False) or hols_df.columns[1]
        he = pick_col(hols_df, ["EndDate", "End"], required=False) or hols_df.columns[2]
        notes_c = pick_col(hols_df, ["Notes", "Note", "Reason"], required=False)
        for _, r in hols_df.iterrows():
            nm = str(r[hn]).strip()
            sd = to_date(r[hs])
            ed = to_date(r[he])
            note = "" if (not notes_c or pd.isna(r[notes_c])) else str(r[notes_c]).strip().lower()
            kind = "Holiday"
            if "sick" in note or "sickness" in note:
                kind = "Sick"
            elif "bank" in note:
                kind = "Bank Holiday"
            hols.append((nm, sd, ed, kind))

    return staff_df, hours_df, hols, {"staff": staff_sheet, "hours": hours_sheet, "holidays": hols_sheet or ""}

# ---------------- Business rules ----------------
DAY_START = time(8, 0)
DAY_END = time(18, 30)
SLOT_MIN = 30

MIN_BLOCK_SLOTS = 4   # 2h
MAX_BLOCK_SLOTS = 6   # 3h
EMAIL_MAX_BLOCK_SLOTS = 8  # 4h (allowed exception)

BREAK_WINDOW = (time(12, 0), time(14, 0))
BREAK_THRESHOLD_HOURS = 6.0

def timeslots():
    cur = datetime(2000, 1, 1, DAY_START.hour, DAY_START.minute)
    end = datetime(2000, 1, 1, DAY_END.hour, DAY_END.minute)
    out = []
    while cur < end:
        out.append(cur.time())
        cur += timedelta(minutes=SLOT_MIN)
    return out

# Mandatory coverage baseline (per slot)
MANDATORY_RULES = [
    ("FrontDesk_SLGP", DAY_START, DAY_END, 1),
    ("FrontDesk_JEN",  DAY_START, DAY_END, 1),
    ("FrontDesk_BGS",  DAY_START, DAY_END, 1),
    ("Triage_Admin_SLGP", DAY_START, time(16, 0), 1),
    ("Triage_Admin_JEN",  DAY_START, time(16, 0), 1),
    ("Email_Box", DAY_START, DAY_END, 1),
    ("Phones", DAY_START, DAY_END, 2),
    ("Bookings", DAY_START, DAY_END, 3),
]

def awaiting_site_for_day(d: date) -> str:
    wd = d.weekday()  # Mon=0
    if wd in (0, 4):  # Mon/Fri
        return "SLGP"
    if wd in (1, 3):  # Tue/Thu
        return "JEN"
    return "BGS"      # Wed

def awaiting_required(d: date, t: time) -> bool:
    return t_in_range(t, time(10, 0), time(16, 0))

def awaiting_optional(d: date, t: time) -> bool:
    return t_in_range(t, time(16, 0), DAY_END)

def filler_order_for_day(d: date):
    # Mon/Fri prioritise phones over EMIS/Docman
    if d.weekday() in (0, 4):
        return ["Phones", "Bookings", "Awaiting_PSA_Admin", "Emis_Tasks", "Docman_Tasks"]
    return ["Phones", "Bookings", "Emis_Tasks", "Docman_Tasks", "Awaiting_PSA_Admin"]

# ---------------- Skills + weights ----------------
def can(sr, col: str) -> bool:
    return bool(sr.get(col, False))

def skill_allowed(sr, role: str) -> bool:
    if role.startswith("FrontDesk"):
        return can(sr, "CanFrontDesk")
    if role.startswith("Triage_Admin"):
        return can(sr, "CanTriage")
    if role == "Email_Box":
        return can(sr, "CanEmail")
    if role == "Phones":
        return can(sr, "CanPhones")
    if role == "Bookings":
        return can(sr, "CanBookings")
    if role == "Emis_Tasks":
        return can(sr, "CanEMIS")
    if role == "Docman_Tasks":
        return can(sr, "CanDocman_PSA") or can(sr, "CanDocman_AWAIT")
    if role == "Awaiting_PSA_Admin":
        return can(sr, "CanDocman_PSA") or can(sr, "CanDocman_AWAIT")
    if role == "Unassigned":
        return True
    return False

def skill_weight(sr, role: str) -> int:
    # Optional per-person tie-break preference (0–5). Missing/blank -> 3.
    mapping = {
        "Phones": "PhonesWeight",
        "Bookings": "BookingsWeight",
        "Emis_Tasks": "EmisWeight",
        "Docman_Tasks": "DocmanWeight",
        "Awaiting_PSA_Admin": "AwaitingWeight",
        "Email_Box": "EmailWeight",
        "Triage_Admin_SLGP": "TriageWeight",
        "Triage_Admin_JEN": "TriageWeight",
        "FrontDesk_SLGP": "FrontDeskWeight",
        "FrontDesk_JEN": "FrontDeskWeight",
        "FrontDesk_BGS": "FrontDeskWeight",
    }
    col = mapping.get(role)
    if not col:
        return 3
    try:
        v = sr.get(col, 3)
        if pd.isna(v):
            return 3
        v = int(float(v))
        return max(0, min(5, v))
    except Exception:
        return 3

def role_site(role: str, d: date) -> str:
    if role.endswith("_SLGP"):
        return "SLGP"
    if role.endswith("_JEN"):
        return "JEN"
    if role.endswith("_BGS"):
        return "BGS"
    if role == "Bookings":
        return "SLGP"
    if role == "Awaiting_PSA_Admin":
        return awaiting_site_for_day(d)
    return ""

def is_frontdesk(role: str) -> bool:
    return role.startswith("FrontDesk_")

def is_triage(role: str) -> bool:
    return role.startswith("Triage_Admin_")

def holiday_kind(name: str, d: date, hols):
    for n, s, e, k in hols:
        if n.strip().lower() == name.strip().lower() and s and e and s <= d <= e:
            return k
    return None

# =========================================================
# Break placement (closest to midpoint, within 12:00–14:00)
# =========================================================
def build_maps(staff_df, hours_df):
    staff_by_name = {r["Name"]: r for _, r in staff_df.iterrows()}
    hmap = {r["Name"]: r for _, r in hours_df.iterrows()}
    return staff_by_name, hmap

def shift_window(hmap, week_start, d, name):
    days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    dname = days[(d - week_start).days]
    hr = hmap.get(name)
    if hr is None:
        return None, None
    return hr.get(f"{dname}Start"), hr.get(f"{dname}End")

def is_working(hmap, week_start, d, t, name):
    stt, end = shift_window(hmap, week_start, d, name)
    return bool(stt and end and (t >= stt) and (t < end))

def pick_break_slot_near_midpoint(d: date, stt: time, end: time):
    if not stt or not end:
        return None
    if (dt_of(d, end) - dt_of(d, stt)).total_seconds() / 3600.0 <= BREAK_THRESHOLD_HOURS:
        return None
    midpoint = dt_of(d, stt) + (dt_of(d, end) - dt_of(d, stt)) / 2
    candidates = []
    for t in [time(12, 0), time(12, 30), time(13, 0), time(13, 30)]:
        if t >= stt and add_minutes(t, 30) <= end and t_in_range(t, BREAK_WINDOW[0], BREAK_WINDOW[1]):
            dist = abs((dt_of(d, t) - midpoint).total_seconds())
            candidates.append((dist, t))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0])
    return candidates[0][1]

# =========================================================
# Block-based scheduler
# =========================================================
def rota_generate_week(staff_df, hours_df, hols, week_start: date):
    slots = timeslots()
    dates = [week_start + timedelta(days=i) for i in range(5)]
    staff_by_name, hmap = build_maps(staff_df, hours_df)

    # Break map: (d,t)->set(names)
    breaks = {}
    for d in dates:
        for name in staff_by_name.keys():
            if holiday_kind(name, d, hols):
                continue
            stt, end = shift_window(hmap, week_start, d, name)
            b = pick_break_slot_near_midpoint(d, stt, end)
            if b:
                breaks.setdefault((d, b), set()).add(name)

    # Assignments: (d,t,name)->role
    a = {}
    # Active blocks: (d,name)->(role, end_idx_excl)
    active = {}
    # Minutes per role: (d,name,role)->minutes
    role_minutes = {}
    gaps = []

    def on_break(name, d, t):
        return name in breaks.get((d, t), set())

    def is_free(name, d, t):
        return (d, t, name) not in a

    def staff_assigned_to(role, d, t):
        return [nm for (dd, tt, nm), rr in a.items() if dd == d and tt == t and rr == role]

    def needed_for_slot(d, t):
        wanted = []
        for role, ta, tb, need in MANDATORY_RULES:
            if t_in_range(t, ta, tb):
                wanted.append((role, need))
        if awaiting_required(d, t) or awaiting_optional(d, t):
            wanted.append(("Awaiting_PSA_Admin", 1))
        return wanted

    def candidate_ok(name, role, d, t):
        sr = staff_by_name[name]

        if holiday_kind(name, d, hols):
            return False
        if not is_working(hmap, week_start, d, t, name):
            return False
        if on_break(name, d, t):
            return False
        if not is_free(name, d, t):
            return False
        if not skill_allowed(sr, role):
            return False

        home = str(sr.get("HomeSite", "")).strip().upper()
        target_site = role_site(role, d)

        # Front desk: strict home-site only
        if is_frontdesk(role):
            return home == target_site

        # Triage: strict home-site only (THIS FIXES CHRISTINE IN SLGP TRIAGE)
        if is_triage(role):
            return home == target_site

        # Bookings: SLGP only
        if role == "Bookings":
            return home == "SLGP"

        # Awaiting: strict to site-of-day (no cross-site)
        if role == "Awaiting_PSA_Admin":
            return home == awaiting_site_for_day(d)

        # Email: JEN/BGS only (no SLGP)
        if role == "Email_Box":
            return home in ("JEN", "BGS")

        # Phones/Emis/Docman: prefer JEN/BGS; allow SLGP only if they have skills (still permitted)
        if role in ("Phones", "Emis_Tasks", "Docman_Tasks"):
            return True

        return True

    def pick_block_len(name, role, d, start_idx):
        stt, end = shift_window(hmap, week_start, d, name)
        if not stt or not end:
            return 0

        # shift end index
        end_idx = start_idx
        while end_idx < len(slots) and slots[end_idx] < end:
            end_idx += 1

        # break cut
        break_idx = None
        for k in range(start_idx, min(len(slots), end_idx)):
            if name in breaks.get((d, slots[k]), set()):
                break_idx = k
                break

        hard_end = break_idx if break_idx is not None else end_idx
        remaining = hard_end - start_idx
        if remaining <= 0:
            return 0

        # allow small remainder at end of shift
        if remaining < MIN_BLOCK_SLOTS:
            return remaining

        if role == "Email_Box":
            return min(EMAIL_MAX_BLOCK_SLOTS, remaining)

        # triage/frontdesk max 3h by rule (MAX_BLOCK_SLOTS = 3h)
        return min(MAX_BLOCK_SLOTS, remaining)

    def start_block(name, role, d, start_idx):
        L = pick_block_len(name, role, d, start_idx)
        if L <= 0:
            return False
        active[(d, name)] = (role, start_idx + L)
        return True

    def apply_block_if_active(name, d, idx):
        b = active.get((d, name))
        if not b:
            return False
        role, end_idx = b
        if idx >= end_idx:
            del active[(d, name)]
            return False
        t = slots[idx]
        a[(d, t, name)] = role
        role_minutes[(d, name, role)] = role_minutes.get((d, name, role), 0) + SLOT_MIN
        return True

    def enforce_frontdesk_exactly_one(d, idx):
        t = slots[idx]
        for site in ("SLGP", "JEN", "BGS"):
            role = f"FrontDesk_{site}"
            current = staff_assigned_to(role, d, t)

            if len(current) > 1:
                # should never happen; mark gap
                gaps.append((d, t, role, f"More than 1 assigned ({', '.join(current)})"))
                # keep the first, unassign others
                for extra in current[1:]:
                    a[(d, t, extra)] = "Unassigned"
                current = current[:1]

            if len(current) == 1:
                continue

            # pick candidate by least minutes on this role, weights as tie-break
            cands = []
            for name, sr in staff_by_name.items():
                if not candidate_ok(name, role, d, t):
                    continue
                used_mins = role_minutes.get((d, name, role), 0)
                primary = used_mins / 60.0
                tie = -(skill_weight(sr, role) - 3)
                cands.append(((primary, tie), name))

            if not cands:
                gaps.append((d, t, role, "No suitable staff available"))
                continue

            cands.sort(key=lambda x: x[0])
            pick = cands[0][1]
            if (d, pick) not in active:
                start_block(pick, role, d, idx)
            apply_block_if_active(pick, d, idx)

    def enforce_role_need(role, need, d, idx):
        t = slots[idx]
        current = staff_assigned_to(role, d, t)
        while len(current) < need:
            cands = []
            for name, sr in staff_by_name.items():
                # If already in an active block, they can only satisfy the current role
                b = active.get((d, name))
                if b is not None and b[0] != role:
                    continue
                if not candidate_ok(name, role, d, t):
                    continue

                used_mins = role_minutes.get((d, name, role), 0)
                primary = used_mins / 60.0
                tie = -(skill_weight(sr, role) - 3)  # tie-break only
                cands.append(((primary, tie), name))

            if not cands:
                # Only log gaps for mandatory window; awaiting optional after 16:00 can be skipped
                if role != "Awaiting_PSA_Admin" or awaiting_required(d, t):
                    gaps.append((d, t, role, "No suitable staff available"))
                break

            cands.sort(key=lambda x: x[0])
            pick = cands[0][1]
            if (d, pick) not in active:
                start_block(pick, role, d, idx)
            apply_block_if_active(pick, d, idx)
            current = staff_assigned_to(role, d, t)

    # Main loop
    for d in dates:
        for idx, t in enumerate(slots):
            # extend existing blocks first
            for name in staff_by_name.keys():
                if (d, t, name) in a:
                    continue
                apply_block_if_active(name, d, idx)

            # enforce front desk per site
            enforce_frontdesk_exactly_one(d, idx)

            # enforce all other mandatory roles
            for role, need in needed_for_slot(d, t):
                if role.startswith("FrontDesk_"):
                    continue
                enforce_role_need(role, need, d, idx)

            # Fill remaining free staff with filler tasks (do not leave working people unassigned)
            filler_roles = filler_order_for_day(d)

            for name, sr in staff_by_name.items():
                if holiday_kind(name, d, hols):
                    continue
                if not is_working(hmap, week_start, d, t, name):
                    continue
                if on_break(name, d, t):
                    continue
                if not is_free(name, d, t):
                    continue

                home = str(sr.get("HomeSite", "")).strip().upper()
                chosen = None

                # Enforce "majority of SLGP on bookings" by strongly biasing SLGP fillers to Bookings
                if home == "SLGP" and skill_allowed(sr, "Bookings"):
                    chosen = "Bookings"

                if not chosen:
                    best = None
                    for role in filler_roles:
                        # Awaiting optional only after 16:00; mandatory is already enforced above
                        if role == "Awaiting_PSA_Admin" and not awaiting_optional(d, t):
                            continue
                        # skip if not allowed/doesn't make sense for this person/site
                        if not skill_allowed(sr, role):
                            continue
                        # Awaiting strict site/day
                        if role == "Awaiting_PSA_Admin" and home != awaiting_site_for_day(d):
                            continue
                        # Email strict JEN/BGS
                        if role == "Email_Box" and home not in ("JEN", "BGS"):
                            continue

                        used = role_minutes.get((d, name, role), 0)
                        # Don't keep someone on one filler all day: cap 6h per filler (except Email which is allowed longer)
                        if role != "Email_Box" and used >= 360:
                            continue
                        primary = used
                        tie = -(skill_weight(sr, role) - 3)
                        score = (primary, tie)
                        if best is None or score < best[0]:
                            best = (score, role)
                    if best:
                        chosen = best[1]

                if not chosen:
                    chosen = "Unassigned"

                if (d, name) not in active:
                    start_block(name, chosen, d, idx)
                apply_block_if_active(name, d, idx)

    # Break sanity check
    for d in dates:
        for name in staff_by_name.keys():
            if holiday_kind(name, d, hols):
                continue
            stt, end = shift_window(hmap, week_start, d, name)
            if not stt or not end:
                continue
            if (dt_of(d, end) - dt_of(d, stt)).total_seconds() / 3600.0 <= BREAK_THRESHOLD_HOURS:
                continue
            had_break = any(name in breaks.get((d, bt), set()) for bt in (time(12,0), time(12,30), time(13,0), time(13,30)))
            if not had_break:
                gaps.append((d, None, "Break", f"{name}: shift > 6h but no break could be placed 12:00–14:00"))

    return a, breaks, gaps, hmap

# =========================================================
# Excel output (colour-coded)
# =========================================================
ROLE_COLORS = {
    "FrontDesk_SLGP": "FFF2CC",
    "FrontDesk_JEN":  "FFF2CC",
    "FrontDesk_BGS":  "FFF2CC",
    "Triage_Admin_SLGP": "D9EAD3",
    "Triage_Admin_JEN":  "D9EAD3",
    "Email_Box": "CFE2F3",
    "Phones": "C9DAF8",
    "Bookings": "FCE5CD",
    "Emis_Tasks": "EAD1DC",
    "Docman_Tasks": "D0E0E3",
    "Awaiting_PSA_Admin": "D0E0E3",
    "Unassigned": "FFFFFF",
    "Break": "CFE2F3",
    "Holiday": "FFF2CC",
    "Bank Holiday": "FFE599",
    "Sick": "F4CCCC",
    "": "DDDDDD",  # non-working blank grey
}

def fill_for(value: str):
    return PatternFill("solid", fgColor=ROLE_COLORS.get(value, "FFFFFF"))

def build_workbook(staff_df, hours_df, hols, start_monday: date, weeks: int):
    wb = Workbook()
    wb.remove(wb.active)
    slots = timeslots()

    for w in range(weeks):
        week_start = start_monday + timedelta(days=7 * w)
        dates = [week_start + timedelta(days=i) for i in range(5)]
        staff_names = list(staff_df["Name"].astype(str))

        a, breaks, gaps, hmap = rota_generate_week(staff_df, hours_df, hols, week_start)

        # Grid
        ws_grid = wb.create_sheet(f"Week{w+1}_Grid")
        ws_grid.append(["Time"] + [d.strftime("%a %d-%b") for d in dates])
        for c in ws_grid[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")

        display_roles = [
            "FrontDesk_SLGP","FrontDesk_JEN","FrontDesk_BGS",
            "Triage_Admin_SLGP","Triage_Admin_JEN",
            "Email_Box","Phones","Bookings",
            "Awaiting_PSA_Admin","Emis_Tasks","Docman_Tasks",
        ]

        for t in slots:
            row = [t.strftime("%H:%M")]
            for d in dates:
                parts = []
                for role in display_roles:
                    ppl = [nm for nm in staff_names if a.get((d, t, nm)) == role]
                    if ppl:
                        parts.append(f"{role}: " + ", ".join(ppl))
                row.append("\n".join(parts))
            ws_grid.append(row)

        ws_grid.column_dimensions["A"].width = 8
        for col in range(2, 7):
            ws_grid.column_dimensions[chr(64 + col)].width = 42
        for r in range(2, 2 + len(slots)):
            ws_grid.row_dimensions[r].height = 95
            for c in range(2, 7):
                ws_grid.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")

        # Coverage gaps
        ws_gaps = wb.create_sheet(f"Week{w+1}_CoverageGaps")
        ws_gaps.append(["Date", "Time", "Role", "Issue"])
        for c in ws_gaps[1]:
            c.font = Font(bold=True)
        for d, t, role, issue in gaps:
            ws_gaps.append([d.isoformat(), "" if t is None else t.strftime("%H:%M"), role, issue])

        # Totals
        ws_tot = wb.create_sheet(f"Week{w+1}_Totals")
        rows = []
        for d in dates:
            for t in slots:
                for nm in staff_names:
                    role = a.get((d, t, nm))
                    if role:
                        rows.append([d, nm, role, 0.5])
                for nm in breaks.get((d, t), set()):
                    rows.append([d, nm, "Break", 0.5])

        df = pd.DataFrame(rows, columns=["Date", "Name", "Task", "Hours"])
        if df.empty:
            df = pd.DataFrame(columns=["Date", "Name", "Task", "Hours"])

        pivot_w = (
            df.groupby(["Name", "Task"])["Hours"].sum()
            .reset_index()
            .pivot(index="Name", columns="Task", values="Hours")
            .fillna(0.0)
        )
        pivot_w["WeeklyTotal"] = pivot_w.sum(axis=1)
        pivot_w = pivot_w.reset_index()

        pivot_d = df.groupby(["Date", "Name", "Task"])["Hours"].sum().reset_index()

        ws_tot.append(["Weekly totals per staff by task (hours)"])
        ws_tot["A1"].font = Font(bold=True, size=12)
        ws_tot.append([])
        for r in dataframe_to_rows(pivot_w, index=False, header=True):
            ws_tot.append(r)

        header_row = 3
        for cell in ws_tot[header_row]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_tot.append([])
        ws_tot.append(["Daily totals per staff by task (hours)"])
        ws_tot[f"A{ws_tot.max_row}"].font = Font(bold=True, size=12)
        ws_tot.append([])
        for r in dataframe_to_rows(pivot_d, index=False, header=True):
            ws_tot.append(r)

        # Timelines (All staff)
        ws_tl = wb.create_sheet(f"Week{w+1}_Timelines_All")
        ws_tl.append(["Date", "Time"] + staff_names)
        for c in ws_tl[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_tl.freeze_panes = "C2"

        for d in dates:
            for t in slots:
                row = [d.strftime("%a %d-%b"), t.strftime("%H:%M")]
                for nm in staff_names:
                    hk = holiday_kind(nm, d, hols)
                    if hk:
                        row.append(hk)
                    elif not is_working(hmap, week_start, d, t, nm):
                        row.append("")  # blank grey
                    elif nm in breaks.get((d, t), set()):
                        row.append("Break")
                    else:
                        row.append(a.get((d, t, nm), "Unassigned"))
                ws_tl.append(row)

        for rr in range(2, ws_tl.max_row + 1):
            for cc in range(3, ws_tl.max_column + 1):
                val = str(ws_tl.cell(rr, cc).value or "")
                ws_tl.cell(rr, cc).fill = fill_for(val)
                ws_tl.cell(rr, cc).alignment = Alignment(wrap_text=True, vertical="top")

        # Timelines by site
        for site in ("SLGP", "JEN", "BGS"):
            site_staff = list(staff_df.loc[staff_df["HomeSite"].astype(str).str.upper() == site, "Name"].astype(str))
            if not site_staff:
                continue
            ws_site = wb.create_sheet(f"Week{w+1}_{site}_Timelines")
            ws_site.append(["Date", "Time"] + site_staff)
            for c in ws_site[1]:
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws_site.freeze_panes = "C2"

            for d in dates:
                for t in slots:
                    row = [d.strftime("%a %d-%b"), t.strftime("%H:%M")]
                    for nm in site_staff:
                        hk = holiday_kind(nm, d, hols)
                        if hk:
                            row.append(hk)
                        elif not is_working(hmap, week_start, d, t, nm):
                            row.append("")  # blank grey
                        elif nm in breaks.get((d, t), set()):
                            row.append("Break")
                        else:
                            row.append(a.get((d, t, nm), "Unassigned"))
                    ws_site.append(row)

            for rr in range(2, ws_site.max_row + 1):
                for cc in range(3, ws_site.max_column + 1):
                    val = str(ws_site.cell(rr, cc).value or "")
                    ws_site.cell(rr, cc).fill = fill_for(val)
                    ws_site.cell(rr, cc).alignment = Alignment(wrap_text=True, vertical="top")

    return wb

# =========================================================
# Streamlit UI
# =========================================================
st.set_page_config(page_title="Rota Generator v10d", layout="wide")
require_password()

st.title("Rota Generator v10d — Clean Rebuild")

st.markdown(
    """
### First time using this rota generator?
1. Download the starter template below
2. Fill in **Staff**, **WorkingHours**, and (optional) **Holidays**
3. Optional: add task **weights (0–5)** per staff member to prefer stronger/faster people for specific tasks
4. Upload the completed template and click **Generate**

**Weights are tie-breakers only** — if weights are blank or 3, behaviour is unchanged.
"""
)

# Template download (must exist in repo root)
tpl_bytes = None
for p in ("rota_generator_template.xlsx", "rota_generator_template_v2.xlsx"):
    try:
        with open(p, "rb") as f:
            tpl_bytes = f.read()
        break
    except Exception:
        continue

if tpl_bytes:
    st.download_button(
        "📥 Download starter template",
        data=tpl_bytes,
        file_name="rota_generator_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.info("Add 'rota_generator_template.xlsx' to your repo root to enable template download.")

uploaded = st.file_uploader("Upload your completed template (.xlsx)", type=["xlsx"])

c1, c2 = st.columns(2)
with c1:
    start_date = st.date_input("Week commencing (Monday)", value=date.today())
with c2:
    mode = st.selectbox("Run period", ["1 week", "4 weeks (month)", "Custom # weeks"])

if mode == "Custom # weeks":
    weeks = int(st.number_input("Number of weeks", min_value=1, max_value=12, value=2, step=1))
elif mode == "4 weeks (month)":
    weeks = 4
else:
    weeks = 1

start_monday = ensure_monday(start_date)

if uploaded:
    try:
        staff_df, hours_df, hols, found = read_template(uploaded.getvalue())
        st.success(f"Template loaded. Sheets: Staff={found['staff']} | Hours={found['hours']} | Holidays={found['holidays'] or 'None'}")
        if st.button("Generate rota and download Excel", type="primary"):
            wb = build_workbook(staff_df, hours_df, hols, start_monday, weeks)
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            out_name = f"rota_{start_monday.isoformat()}_{weeks}w.xlsx"
            st.download_button(
                "📊 Download Excel rota",
                data=bio.getvalue(),
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    except Exception as e:
        st.error("Could not process the template.")
        st.exception(e)
else:
    st.info("Upload your completed template to generate the rota.")
