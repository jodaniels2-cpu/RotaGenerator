import io
import re
from datetime import datetime, date, time, timedelta

import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


# =========================================================
# Password protection (Streamlit Secrets)
# =========================================================
def require_password():
    """
    Add to Streamlit secrets (TOML):

    APP_PASSWORD = "your-strong-password"
    """
    pw = st.secrets.get("APP_PASSWORD", None)
    if not pw:
        st.warning("Password is not configured. Add APP_PASSWORD in Streamlit Secrets to enable protection.")
        return True

    if "authed" not in st.session_state:
        st.session_state.authed = False

    if st.session_state.authed:
        return True

    with st.form("login"):
        entered = st.text_input("Password", type="password")
        ok = st.form_submit_button("Log in")
        if ok:
            if entered == pw:
                st.session_state.authed = True
                st.success("Logged in.")
                return True
            st.error("Incorrect password.")
    st.stop()


# =========================================================
# Helpers: robust sheet + column detection
# =========================================================
def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s).strip().lower())

def find_sheet(xls: pd.ExcelFile, candidates):
    names = {normalize(n): n for n in xls.sheet_names}
    for c in candidates:
        key = normalize(c)
        if key in names:
            return names[key]
    # fuzzy contains
    for n in xls.sheet_names:
        nn = normalize(n)
        for c in candidates:
            if normalize(c) in nn or nn in normalize(c):
                return n
    return None

def pick_col(df: pd.DataFrame, candidates, required=True):
    cols = {normalize(c): c for c in df.columns}
    for cand in candidates:
        key = normalize(cand)
        if key in cols:
            return cols[key]
    # fuzzy contains
    for c in df.columns:
        nc = normalize(c)
        for cand in candidates:
            if normalize(cand) in nc:
                return c
    if required:
        raise KeyError(
            f"Could not find required column among: {candidates}. "
            f"Available: {list(df.columns)}"
        )
    return None

def to_time(x):
    if pd.isna(x) or x == "":
        return None
    if isinstance(x, time):
        return x
    if isinstance(x, datetime):
        return x.time()
    if isinstance(x, (float, int)):
        # Excel time as fraction of day
        seconds = int(round(float(x) * 86400))
        return (datetime(2000, 1, 1) + timedelta(seconds=seconds)).time()
    s = str(x).strip()
    for fmt in ("%H:%M", "%H.%M", "%I:%M%p", "%I:%M %p"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            pass
    raise ValueError(f"Unrecognized time format: {x}")

def to_date(x):
    if pd.isna(x) or x == "":
        return None
    if isinstance(x, (datetime, pd.Timestamp)):
        return x.date()
    if isinstance(x, date):
        return x
    return pd.to_datetime(x).date()


# =========================================================
# Parse template (robust)
# =========================================================
def read_template(uploaded_bytes: bytes):
    xls = pd.ExcelFile(io.BytesIO(uploaded_bytes))

    staff_sheet = find_sheet(xls, ["Staff", "Skills", "People"])
    hours_sheet = find_sheet(xls, ["WorkingHours", "Hours", "Availability"])
    hols_sheet = find_sheet(xls, ["Holidays", "Leave", "Absence"])
    params_sheet = find_sheet(xls, ["Parameters", "Params", "Rules", "Config"])

    if not staff_sheet:
        raise ValueError(f"Could not find Staff/Skills sheet. Found: {xls.sheet_names}")
    if not hours_sheet:
        raise ValueError(f"Could not find WorkingHours/Hours sheet. Found: {xls.sheet_names}")
    if not params_sheet:
        raise ValueError(f"Could not find Parameters/Params sheet. Found: {xls.sheet_names}")

    staff_df = pd.read_excel(xls, sheet_name=staff_sheet)
    hours_df = pd.read_excel(xls, sheet_name=hours_sheet)
    hols_df = pd.read_excel(xls, sheet_name=hols_sheet) if hols_sheet else pd.DataFrame()
    params_df = pd.read_excel(xls, sheet_name=params_sheet)

    # --- staff cols ---
    name_c = pick_col(staff_df, ["Name", "StaffName"])
    home_c = pick_col(staff_df, ["HomeSite", "Site", "BaseSite"], required=False)

    staff_df = staff_df.copy()
    staff_df["Name"] = staff_df[name_c].astype(str).str.strip()
    staff_df["HomeSite"] = staff_df[home_c].astype(str).str.strip().fillna("") if home_c else ""

    # normalize Y/N flags: any column that looks like Can* or skill-ish
    def yn(v):
        if pd.isna(v):
            return False
        s = str(v).strip().lower()
        return s in ["y", "yes", "true", "1"]

    # Convert all non-core columns to bool if they contain Y/N
    core_cols_norm = {normalize(name_c), normalize(home_c or ""), "name", "homesite"}
    for c in list(staff_df.columns):
        if normalize(c) in core_cols_norm:
            continue
        # only coerce if it contains any Y/N-ish values
        if staff_df[c].astype(str).str.strip().str.lower().isin(["y","yes","n","no","true","false","1","0"]).any():
            staff_df[c] = staff_df[c].apply(yn)

    # detect Carol Church column (optional)
    carol_c = None
    for c in staff_df.columns:
        if normalize(c) in ["iscarolchurchyn", "iscarolchurch", "carolchurch", "carol"]:
            carol_c = c
            break
    if carol_c:
        staff_df["IsCarolChurch"] = staff_df[carol_c].apply(bool)
    else:
        staff_df["IsCarolChurch"] = staff_df["Name"].str.lower().eq("carol church")

    # --- working hours ---
    hours_df = hours_df.copy()
    hours_name_c = pick_col(hours_df, ["Name", "StaffName"])
    hours_df["Name"] = hours_df[hours_name_c].astype(str).str.strip()

    # day columns: MonStart MonEnd ...
    days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    for d in days:
        sc = pick_col(hours_df, [f"{d}Start", f"{d} Start", f"{d}_Start"], required=False)
        ec = pick_col(hours_df, [f"{d}End", f"{d} End", f"{d}_End"], required=False)
        hours_df[f"{d}Start"] = hours_df[sc].apply(to_time) if sc else None
        hours_df[f"{d}End"] = hours_df[ec].apply(to_time) if ec else None

    # --- holidays ---
    hols = []
    if not hols_df.empty:
        hname = pick_col(hols_df, ["Name", "StaffName"], required=False)
        hs = pick_col(hols_df, ["StartDate", "Start"], required=False)
        he = pick_col(hols_df, ["EndDate", "End"], required=False)
        if hname and hs and he:
            for _, r in hols_df.iterrows():
                hols.append((str(r[hname]).strip(), to_date(r[hs]), to_date(r[he])))

    # --- parameters ---
    rule_c = pick_col(params_df, ["Rule", "Parameter", "Key", "Setting"], required=False)
    val_c = pick_col(params_df, ["Value", "Val"], required=False)
    params = {}
    if rule_c and val_c:
        for _, r in params_df.iterrows():
            k = str(r[rule_c]).strip()
            if k and k.lower() != "nan":
                params[k] = r[val_c]
    else:
        if params_df.shape[1] >= 2:
            c1, c2 = params_df.columns[:2]
            for _, r in params_df.iterrows():
                k = str(r[c1]).strip()
                if k and k.lower() != "nan":
                    params[k] = r[c2]

    return staff_df, hours_df, hols, params, xls.sheet_names


# =========================================================
# Core rota model
# =========================================================
SITES = ["SLGP", "JEN", "BGS"]

ROLE_COLORS = {
    "FrontDesk_SLGP": "FFF2CC",
    "FrontDesk_JEN":  "FFF2CC",
    "FrontDesk_BGS":  "FFF2CC",
    "Triage_Admin_SLGP": "D9EAD3",
    "Triage_Admin_JEN":  "D9EAD3",
    "Email_Box": "CFE2F3",
    "Phones": "C9DAF8",
    "Bookings": "FCE5CD",
    "EMIS": "EAD1DC",
    "Docman_PSA": "D0E0E3",
    "Docman_Awaiting": "D0E0E3",
    "Break": "E6E6E6",
    "Unassigned": "FFFFFF",
}

MANDATORY = [
    ("FrontDesk_SLGP", time(8, 0), time(18, 30), 1),
    ("FrontDesk_JEN",  time(8, 0), time(18, 30), 1),
    ("FrontDesk_BGS",  time(8, 0), time(18, 30), 1),
    ("Triage_Admin_SLGP", time(8, 0), time(16, 0), 1),
    ("Triage_Admin_JEN",  time(8, 0), time(16, 0), 1),
    ("Email_Box", time(8, 0), time(18, 30), 1),
    ("Phones", time(8, 0), time(18, 30), 2),  # min 2, optional 3
]

def timeslots(day_start=time(8, 0), day_end=time(18, 30), step_min=30):
    start_dt = datetime(2000, 1, 1, day_start.hour, day_start.minute)
    end_dt = datetime(2000, 1, 1, day_end.hour, day_end.minute)
    cur = start_dt
    out = []
    while cur < end_dt:
        out.append(cur.time())
        cur += timedelta(minutes=step_min)
    return out

def t_in_range(t, a, b):
    return (t >= a) and (t < b)

def hours_between(t1, t2):
    dt1 = datetime(2000, 1, 1, t1.hour, t1.minute)
    dt2 = datetime(2000, 1, 1, t2.hour, t2.minute)
    return (dt2 - dt1).total_seconds() / 3600

def is_on_holiday(name, d, hols):
    n0 = name.strip().lower()
    for n, s, e in hols:
        if n.strip().lower() == n0 and s and e and s <= d <= e:
            return True
    return False

def skill_allowed(staff_row: pd.Series, role: str) -> bool:
    """
    Matches template headers like:
      CanFrontDesk, CanTriage, CanEmail, CanPhones, CanBookings, CanEMIS,
      CanDocman_PSA, CanDocman_AWAIT
    plus some older variants.
    """
    # Map role -> acceptable column name variants
    role_map = {
        "FrontDesk": ["CanFrontDesk", "FrontDesk", "Front Desk", "Reception"],
        "Triage": ["CanTriage", "Triage", "AdminTriage", "Admin Triage"],
        "Email_Box": ["CanEmail", "Email", "EmailBox", "Emails", "Email Box"],
        "Phones": ["CanPhones", "Phones", "Phone"],
        "Bookings": ["CanBookings", "Bookings", "Booking"],
        "EMIS": ["CanEMIS", "EMIS"],
        "Docman_PSA": ["CanDocman_PSA", "DocmanPSA", "Docman PSA", "PSA"],
        "Docman_Awaiting": ["CanDocman_AWAIT", "CanDocman_Awaiting", "DocmanAwaiting", "AwaitingResponse", "Awaiting Response"],
    }

    if role.startswith("FrontDesk"):
        keys = role_map["FrontDesk"]
    elif role.startswith("Triage_Admin"):
        keys = role_map["Triage"]
    else:
        keys = role_map.get(role, [role])

    cols = {normalize(c): c for c in staff_row.index}
    for k in keys:
        nk = normalize(k)
        if nk in cols:
            return bool(staff_row[cols[nk]])
    return False  # safest default

def site_restriction_ok(staff_row: pd.Series, role: str) -> bool:
    home = str(staff_row.get("HomeSite", "")).strip().upper()

    # Email + Phones + EMIS + Docman only JEN/BGS
    if role in ["Email_Box", "Phones", "EMIS", "Docman_PSA", "Docman_Awaiting"]:
        return home in ["JEN", "BGS"]

    # Bookings only SLGP
    if role == "Bookings":
        return home == "SLGP"

    # site-specific roles
    if role.endswith("_SLGP"):
        return home == "SLGP"
    if role.endswith("_JEN"):
        return home == "JEN"
    if role.endswith("_BGS"):
        return home == "BGS"

    return True

def build_availability(staff_df, hours_df, hols, week_start: date):
    hmap = {r["Name"]: r for _, r in hours_df.iterrows()}
    days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    out = {}
    for i, dname in enumerate(days):
        d = week_start + timedelta(days=i)
        available = []
        for _, s in staff_df.iterrows():
            name = s["Name"]
            if is_on_holiday(name, d, hols):
                continue
            hr = hmap.get(name)
            if hr is None:
                continue
            stt = hr.get(f"{dname}Start")
            end = hr.get(f"{dname}End")
            if stt and end:
                available.append(name)
        out[d] = available
    return out, hmap

def staff_work_window(hmap_row, dname):
    stt = hmap_row.get(f"{dname}Start")
    end = hmap_row.get(f"{dname}End")
    return stt, end


# =========================================================
# Scheduler
# =========================================================
def rota_generate_one_week(
    staff_df,
    hours_df,
    hols,
    week_start: date,
    phones_max=3,
    break_window=(time(11, 30), time(14, 0)),
    max_frontdesk_block_hours=2.5,
    max_triage_block_hours=3.0,
    fairness_state=None,
):
    """
    Returns:
      assign_mandatory: dict[(date,time)] -> dict[role_key] = staffname
      staff_timeline: dict[(date,time,staff)] -> role string (includes filler tasks + Break)
      gaps: list tuples (date, time|None, role, issue)
      fairness_state: carry across weeks
    """
    if fairness_state is None:
        fairness_state = {"role_hours": {}, "frontdesk_hours": {}, "triage_hours": {}}

    slots = timeslots()
    days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    dates = [week_start + timedelta(days=i) for i in range(5)]

    availability, hmap = build_availability(staff_df, hours_df, hols, week_start)
    staff_by_name = {r["Name"]: r for _, r in staff_df.iterrows()}

    assign = {}  # mandatory coverage only
    gaps = []

    # role blocks
    frontdesk_block = {}
    triage_block = {}

    # per-staff daily assigned (mandatory) hours
    day_mandatory_hours = {}

    # planned breaks: (date, staff)-> break_time_slot
    break_slot = {}

    def can_work(name, d, t):
        hr = hmap.get(name)
        if hr is None:
            return False
        dname = days[(d - week_start).days]
        stt, end = staff_work_window(hr, dname)
        return bool(stt and end and (t >= stt) and (t < end))

    def score_candidate(name, role):
        key = (name, role)
        base = fairness_state["role_hours"].get(key, 0.0)
        if role.startswith("FrontDesk"):
            base += 2.0 * fairness_state["frontdesk_hours"].get(name, 0.0)
        if role.startswith("Triage_Admin"):
            base += 1.5 * fairness_state["triage_hours"].get(name, 0.0)
        # prefer people with fewer mandatory hours today
        base += 0.3 * day_mandatory_hours.get((role, name), 0.0)
        return base

    def pick_staff(d, t, role, already_used):
        # Carol rule override: if Carol working, she is on front desk at her home site
        if role.startswith("FrontDesk"):
            for _, sr in staff_df.iterrows():
                if sr.get("IsCarolChurch", False) and can_work(sr["Name"], d, t):
                    home = str(sr.get("HomeSite", "")).upper()
                    if role.endswith(home):
                        return sr["Name"]

        candidates = []
        for name in availability.get(d, []):
            if name in already_used:
                continue
            if not can_work(name, d, t):
                continue
            sr = staff_by_name[name]
            if not site_restriction_ok(sr, role):
                continue
            if role.startswith("FrontDesk"):
                if not skill_allowed(sr, "FrontDesk"):
                    continue
                cur = frontdesk_block.get((d, name), 0.0)
                if cur + 0.5 > max_frontdesk_block_hours:
                    continue
            elif role.startswith("Triage_Admin"):
                if not skill_allowed(sr, "Triage"):
                    continue
                cur = triage_block.get((d, name), 0.0)
                if cur + 0.5 > max_triage_block_hours:
                    continue
            else:
                if not skill_allowed(sr, role):
                    continue

            candidates.append(name)

        if not candidates:
            return None
        candidates.sort(key=lambda n: score_candidate(n, role))
        return candidates[0]

    # Pre-plan breaks for anyone with shift > 6h (choose first free slot in window)
    bw0, bw1 = break_window
    for d in dates:
        dname = days[(d - week_start).days]
        for name in availability.get(d, []):
            hr = hmap[name]
            stt, end = staff_work_window(hr, dname)
            if not stt or not end:
                continue
            if hours_between(stt, end) <= 6:
                continue

            # pick first workable slot in break window
            chosen = None
            for t in slots:
                if not t_in_range(t, bw0, bw1):
                    continue
                if not can_work(name, d, t):
                    continue
                chosen = t
                break
            if chosen:
                break_slot[(d, name)] = chosen
            else:
                gaps.append((d, None, "Break", f"{name} shift > 6h but no possible slot in 11:30–14:00"))

    # Assign mandatory roles slot-by-slot
    for d in dates:
        for t in slots:
            used = set()
            slot_roles = assign.setdefault((d, t), {})

            for role, r0, r1, need in MANDATORY:
                if not t_in_range(t, r0, r1):
                    continue

                if role == "Phones":
                    # enforce min 2
                    for k in range(need):
                        pick = pick_staff(d, t, role, used)
                        if pick:
                            slot_roles[f"Phones_{k+1}"] = pick
                            used.add(pick)
                            day_mandatory_hours[(role, pick)] = day_mandatory_hours.get((role, pick), 0.0) + 0.5
                        else:
                            gaps.append((d, t, "Phones", "No available JEN/BGS phone-trained staff"))

                    # optional 3rd
                    if phones_max >= 3:
                        pick = pick_staff(d, t, role, used)
                        if pick:
                            slot_roles["Phones_3"] = pick
                            used.add(pick)
                            day_mandatory_hours[(role, pick)] = day_mandatory_hours.get((role, pick), 0.0) + 0.5
                    continue

                pick = pick_staff(d, t, role, used)
                if pick:
                    slot_roles[role] = pick
                    used.add(pick)
                    day_mandatory_hours[(role, pick)] = day_mandatory_hours.get((role, pick), 0.0) + 0.5

                    if role.startswith("FrontDesk"):
                        frontdesk_block[(d, pick)] = frontdesk_block.get((d, pick), 0.0) + 0.5
                    if role.startswith("Triage_Admin"):
                        triage_block[(d, pick)] = triage_block.get((d, pick), 0.0) + 0.5
                else:
                    gaps.append((d, t, role, "No suitable staff available"))

            # reset blocks for staff not on those roles this slot
            fd_names = {v for k, v in slot_roles.items() if k.startswith("FrontDesk")}
            tr_names = {v for k, v in slot_roles.items() if k.startswith("Triage_Admin")}
            for name in availability.get(d, []):
                if name not in fd_names:
                    frontdesk_block[(d, name)] = 0.0
                if name not in tr_names:
                    triage_block[(d, name)] = 0.0

    # Build per-staff per-slot timeline including filler tasks
    staff_timeline = {}  # (d,t,name)->role
    # quick lookup: slot -> who is on what mandatory role
    slot_person_role = {}  # (d,t,name)->role
    for (d, t), slot_roles in assign.items():
        for role_key, nm in slot_roles.items():
            role = "Phones" if role_key.startswith("Phones_") else role_key
            slot_person_role[(d, t, nm)] = role

    # Filler task pick (deterministic + "majority of SLGP on bookings")
    def filler_role_for_staff(sr: pd.Series):
        home = str(sr.get("HomeSite", "")).strip().upper()
        if home == "SLGP":
            if skill_allowed(sr, "Bookings"):
                return "Bookings"
            return "Unassigned"

        # JEN/BGS priorities
        if skill_allowed(sr, "EMIS"):
            return "EMIS"
        if skill_allowed(sr, "Docman_PSA"):
            return "Docman_PSA"
        if skill_allowed(sr, "Docman_Awaiting"):
            return "Docman_Awaiting"
        return "Unassigned"

    # Create timeline: if staff working and not mandatory, assign filler, except break slot
    for d in dates:
        dname = days[(d - week_start).days]
        for name in availability.get(d, []):
            sr = staff_by_name[name]
            for t in slots:
                if not can_work(name, d, t):
                    continue

                # break wins if planned and staff not holding mandatory at that time
                if break_slot.get((d, name)) == t and (d, t, name) not in slot_person_role:
                    staff_timeline[(d, t, name)] = "Break"
                    continue

                if (d, t, name) in slot_person_role:
                    staff_timeline[(d, t, name)] = slot_person_role[(d, t, name)]
                else:
                    staff_timeline[(d, t, name)] = filler_role_for_staff(sr)

    # Enforce "majority of SLGP on bookings" mathematically (per day):
    # For SLGP staff with Bookings skill, any non-mandatory non-break slot must be Bookings.
    for d in dates:
        for name in availability.get(d, []):
            sr = staff_by_name[name]
            home = str(sr.get("HomeSite", "")).strip().upper()
            if home != "SLGP":
                continue
            if not skill_allowed(sr, "Bookings"):
                continue
            for t in slots:
                key = (d, t, name)
                if key not in staff_timeline:
                    continue
                if staff_timeline[key] in ["Break", "FrontDesk_SLGP", "Triage_Admin_SLGP"]:
                    continue
                # if it's anything else, force Bookings
                staff_timeline[key] = "Bookings"

    # Update fairness_state using timeline (hours by role)
    for (d, t, name), role in staff_timeline.items():
        if role == "Unassigned":
            continue
        fairness_state["role_hours"][(name, role)] = fairness_state["role_hours"].get((name, role), 0.0) + 0.5
        if role.startswith("FrontDesk"):
            fairness_state["frontdesk_hours"][name] = fairness_state["frontdesk_hours"].get(name, 0.0) + 0.5
        if role.startswith("Triage_Admin"):
            fairness_state["triage_hours"][name] = fairness_state["triage_hours"].get(name, 0.0) + 0.5

    # Break compliance check: if shift > 6, ensure at least one Break slot in timeline
    for d in dates:
        dname = days[(d - week_start).days]
        for name in availability.get(d, []):
            hr = hmap[name]
            stt, end = staff_work_window(hr, dname)
            if not stt or not end:
                continue
            if hours_between(stt, end) <= 6:
                continue
            has_break = any(
                staff_timeline.get((d, t, name)) == "Break"
                for t in slots
                if can_work(name, d, t)
            )
            if not has_break:
                gaps.append((d, None, "Break", f"{name} shift > 6h but no break assigned"))

    return assign, staff_timeline, gaps, fairness_state


# =========================================================
# Excel writer
# =========================================================
def fill_for_role(role):
    col = ROLE_COLORS.get(role, "FFFFFF")
    return PatternFill("solid", fgColor=col)

def write_week_to_workbook(wb: Workbook, title: str, week_start: date, assign, staff_timeline, gaps):
    slots = timeslots()
    days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    dates = [week_start + timedelta(days=i) for i in range(5)]

    ws_grid = wb.create_sheet(f"{title}_Coverage")
    ws_tl = wb.create_sheet(f"{title}_StaffTimelines")
    ws_gap = wb.create_sheet(f"{title}_CoverageGaps")
    ws_tot = wb.create_sheet(f"{title}_Totals")

    # -----------------
    # Coverage grid (mandatory) by time slot
    # -----------------
    ws_grid.append(["Time"] + [d.strftime("%a %d-%b") for d in dates])
    for cell in ws_grid[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for t in slots:
        row = [t.strftime("%H:%M")]
        for d in dates:
            slot_roles = assign.get((d, t), {})
            parts = []
            for k in ["FrontDesk_SLGP", "FrontDesk_JEN", "FrontDesk_BGS",
                      "Triage_Admin_SLGP", "Triage_Admin_JEN", "Email_Box"]:
                if k in slot_roles:
                    parts.append(f"{k}: {slot_roles[k]}")

            phones = [slot_roles.get("Phones_1"), slot_roles.get("Phones_2"), slot_roles.get("Phones_3")]
            phones = [p for p in phones if p]
            if phones:
                parts.append("Phones: " + ", ".join(phones))

            row.append("\n".join(parts))
        ws_grid.append(row)

    ws_grid.column_dimensions["A"].width = 8
    for col in range(2, 7):
        ws_grid.column_dimensions[chr(64 + col)].width = 34
    for r in range(2, 2 + len(slots)):
        ws_grid.row_dimensions[r].height = 70
        for c in range(2, 7):
            ws_grid.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")

    # -----------------
    # Staff timelines (audit-friendly)
    # One tab per week, grid: rows = staff, cols = day/time
    # -----------------
    # Build ordered staff list from timeline keys
    staff_names = sorted({name for (_, _, name) in staff_timeline.keys()})
    # Header: Day + time slots blocks
    header = ["Name"]
    for d in dates:
        for t in slots:
            header.append(f"{d.strftime('%a')}\n{t.strftime('%H:%M')}")
    ws_tl.append(header)
    for cell in ws_tl[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for name in staff_names:
        row = [name]
        for d in dates:
            for t in slots:
                role = staff_timeline.get((d, t, name), "")
                row.append(role)
        ws_tl.append(row)

    ws_tl.freeze_panes = "B2"
    ws_tl.column_dimensions["A"].width = 22
    for j in range(2, len(header) + 1):
        ws_tl.column_dimensions[ws_tl.cell(1, j).column_letter].width = 10

    # Color-code by role
    for r in range(2, 2 + len(staff_names)):
        ws_tl.row_dimensions[r].height = 18
        for j in range(2, len(header) + 1):
            role = ws_tl.cell(r, j).value
            if role:
                ws_tl.cell(r, j).fill = fill_for_role(role)
            ws_tl.cell(r, j).alignment = Alignment(horizontal="center", vertical="center")

    # -----------------
    # Coverage gaps
    # -----------------
    ws_gap.append(["Date", "Time", "Role", "Issue"])
    for cell in ws_gap[1]:
        cell.font = Font(bold=True)

    for d, t, role, issue in gaps:
        t_str = "" if t is None else t.strftime("%H:%M")
        ws_gap.append([d.isoformat(), t_str, role, issue])

    # -----------------
    # Totals: daily + weekly totals per staff by task; weekly totals per site
    # -----------------
    # Build longform from staff_timeline
    rows = []
    for (d, t, name), role in staff_timeline.items():
        rows.append([d, name, role, 0.5])

    df = pd.DataFrame(rows, columns=["Date", "Name", "Role", "Hours"])
    if df.empty:
        df = pd.DataFrame(columns=["Date", "Name", "Role", "Hours"])

    # weekly totals per staff by task
    pivot_week = (
        df.groupby(["Name", "Role"])["Hours"].sum()
        .reset_index()
        .pivot(index="Name", columns="Role", values="Hours")
        .fillna(0.0)
    )
    pivot_week["WeeklyTotal"] = pivot_week.sum(axis=1)
    pivot_week = pivot_week.reset_index()

    # daily totals per staff by task
    pivot_day = df.groupby(["Date", "Name", "Role"])["Hours"].sum().reset_index()

    # weekly hours totals per site (based on staff HomeSite would be better, but not in timeline here;
    # so we infer from role where possible + otherwise leave in "Other")
    site_map = {
        "FrontDesk_SLGP": "SLGP",
        "Triage_Admin_SLGP": "SLGP",
        "Bookings": "SLGP",
        "FrontDesk_JEN": "JEN",
        "Triage_Admin_JEN": "JEN",
        "EMIS": "JEN/BGS",
        "Docman_PSA": "JEN/BGS",
        "Docman_Awaiting": "JEN/BGS",
        "FrontDesk_BGS": "BGS",
        "Email_Box": "JEN/BGS",
        "Phones": "JEN/BGS",
        "Break": "N/A",
        "Unassigned": "N/A",
    }
    df["SiteBucket"] = df["Role"].map(site_map).fillna("Other")
    site_sum = df.groupby("SiteBucket")["Hours"].sum().reset_index().rename(columns={"Hours": "WeeklyHoursTotal"})

    ws_tot.append(["Weekly totals per staff by task (hours)"])
    ws_tot["A1"].font = Font(bold=True, size=12)

    ws_tot.append([])
    for r in dataframe_to_rows(pivot_week, index=False, header=True):
        ws_tot.append(r)

    # style the pivot header row (row 3)
    header_row = 3
    for cell in ws_tot[header_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # colour columns (role headers)
    headers = [c.value for c in ws_tot[header_row]]
    for j, h in enumerate(headers, start=1):
        if not h or h == "Name":
            continue
        fill = fill_for_role(h)
        for i in range(header_row, header_row + len(pivot_week) + 1):
            ws_tot.cell(i, j).fill = fill

    # daily totals block
    ws_tot.append([])
    ws_tot.append(["Daily totals per staff by task"])
    ws_tot[f"A{ws_tot.max_row}"].font = Font(bold=True, size=12)

    for r in dataframe_to_rows(pivot_day, index=False, header=True):
        ws_tot.append(r)

    # site totals block
    ws_tot.append([])
    ws_tot.append(["Weekly hours totals per site (bucketed)"])
    ws_tot[f"A{ws_tot.max_row}"].font = Font(bold=True, size=12)

    for r in dataframe_to_rows(site_sum, index=False, header=True):
        ws_tot.append(r)

    return wb


def build_excel_for_period(staff_df, hours_df, hols, start_monday: date, weeks: int, phones_max=3):
    wb = Workbook()
    wb.remove(wb.active)

    fairness_state = None
    for w in range(weeks):
        ws = start_monday + timedelta(days=7 * w)
        assign, staff_tl, gaps, fairness_state = rota_generate_one_week(
            staff_df=staff_df,
            hours_df=hours_df,
            hols=hols,
            week_start=ws,
            phones_max=phones_max,
            fairness_state=fairness_state,
        )
        title = f"Week{w+1}_{ws.strftime('%d%b')}"
        write_week_to_workbook(wb, title, ws, assign, staff_tl, gaps)

    return wb


# =========================================================
# Streamlit UI
# =========================================================
st.set_page_config(page_title="Rota Generator", layout="wide")
require_password()

st.title("Rota Generator (Excel export)")

uploaded = st.file_uploader("Upload rota template (.xlsx)", type=["xlsx"])

c1, c2, c3 = st.columns(3)
with c1:
    start_date = st.date_input("Week commencing (Monday)", value=date.today())
with c2:
    mode = st.selectbox("Run period", ["1 week", "4 weeks (month)", "Custom # weeks"])
with c3:
    phones_max = st.selectbox("Max phones if spare staff", [2, 3], index=1)

if mode == "Custom # weeks":
    weeks = st.number_input("Number of weeks", min_value=1, max_value=12, value=2, step=1)
elif mode == "4 weeks (month)":
    weeks = 4
else:
    weeks = 1

def ensure_monday(d: date) -> date:
    return d - timedelta(days=d.weekday())

if uploaded:
    try:
        staff_df, hours_df, hols, params, found_sheets = read_template(uploaded.getvalue())

        start_monday = ensure_monday(start_date)
        st.caption(f"Generating from Monday {start_monday.isoformat()} for {int(weeks)} week(s).")

        if st.button("Generate rota and export Excel"):
            wb = build_excel_for_period(
                staff_df=staff_df,
                hours_df=hours_df,
                hols=hols,
                start_monday=start_monday,
                weeks=int(weeks),
                phones_max=int(phones_max),
            )

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)

            out_name = f"rota_{start_monday.isoformat()}_{int(weeks)}w.xlsx"
            st.download_button(
                "📊 Download Excel rota",
                data=bio.getvalue(),
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.success("Done — download the Excel file above.")

    except Exception as e:
        st.error("Could not process the template.")
        st.exception(e)
else:
    st.info("Upload your completed template to continue.")
