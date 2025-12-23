import io
import re
from datetime import datetime, date, time, timedelta

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================================================
# Password protection (Streamlit Secrets TOML)
# =========================================================
def require_password():
    """
    Streamlit Secrets (TOML):

    APP_PASSWORD = "your-strong-password"
    """
    pw = st.secrets.get("APP_PASSWORD", None)
    if not pw:
        st.warning("APP_PASSWORD not set in Streamlit Secrets. App is not password-protected.")
        return True

    if "authed" not in st.session_state:
        st.session_state.authed = False
    if st.session_state.authed:
        return True

    with st.form("login"):
        entered = st.text_input("Password", type="password")
        ok = st.form_submit_button("Log in")
        if ok:
            if entered == pw:
                st.session_state.authed = True
                st.success("Logged in.")
                return True
            st.error("Incorrect password.")
    st.stop()


# =========================================================
# Helpers
# =========================================================
def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s).strip().lower())

def find_sheet(xls: pd.ExcelFile, candidates):
    names = {normalize(n): n for n in xls.sheet_names}
    for c in candidates:
        key = normalize(c)
        if key in names:
            return names[key]
    for n in xls.sheet_names:
        nn = normalize(n)
        for c in candidates:
            if normalize(c) in nn or nn in normalize(c):
                return n
    return None

def pick_col(df: pd.DataFrame, candidates, required=True):
    cols = {normalize(c): c for c in df.columns}
    for cand in candidates:
        key = normalize(cand)
        if key in cols:
            return cols[key]
    for c in df.columns:
        nc = normalize(c)
        for cand in candidates:
            if normalize(cand) in nc:
                return c
    if required:
        raise KeyError(f"Missing required column among {candidates}. Available: {list(df.columns)}")
    return None

def to_time(x):
    if pd.isna(x) or x == "":
        return None
    if isinstance(x, time):
        return x
    if isinstance(x, datetime):
        return x.time()
    if isinstance(x, (float, int)):
        seconds = int(round(float(x) * 86400))
        return (datetime(2000, 1, 1) + timedelta(seconds=seconds)).time()
    s = str(x).strip()
    for fmt in ("%H:%M", "%H.%M", "%I:%M%p", "%I:%M %p"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            pass
    return pd.to_datetime(s).time()

def to_date(x):
    if pd.isna(x) or x == "":
        return None
    if isinstance(x, (datetime, pd.Timestamp)):
        return x.date()
    if isinstance(x, date):
        return x
    return pd.to_datetime(x).date()

def dt_of(d: date, t: time) -> datetime:
    return datetime(d.year, d.month, d.day, t.hour, t.minute)

def add_minutes(t: time, mins: int) -> time:
    return (datetime(2000, 1, 1, t.hour, t.minute) + timedelta(minutes=mins)).time()

def t_in_range(t: time, a: time, b: time) -> bool:
    return (t >= a) and (t < b)

def h_between(t1: time, t2: time) -> float:
    return (dt_of(date(2000,1,1), t2) - dt_of(date(2000,1,1), t1)).total_seconds() / 3600.0


# =========================================================
# Read template
# =========================================================
def read_template(uploaded_bytes: bytes):
    xls = pd.ExcelFile(io.BytesIO(uploaded_bytes))

    staff_sheet = find_sheet(xls, ["Staff", "Skills", "People"])
    hours_sheet = find_sheet(xls, ["WorkingHours", "Hours", "Availability"])
    hols_sheet = find_sheet(xls, ["Holidays", "Leave", "Absence"])

    if not staff_sheet:
        raise ValueError(f"Could not find Staff sheet. Found: {xls.sheet_names}")
    if not hours_sheet:
        raise ValueError(f"Could not find WorkingHours sheet. Found: {xls.sheet_names}")

    staff_df = pd.read_excel(xls, sheet_name=staff_sheet)
    hours_df = pd.read_excel(xls, sheet_name=hours_sheet)
    hols_df = pd.read_excel(xls, sheet_name=hols_sheet) if hols_sheet else pd.DataFrame()

    # Staff columns
    name_c = pick_col(staff_df, ["Name", "StaffName"])
    home_c = pick_col(staff_df, ["HomeSite", "Site", "BaseSite"], required=False)

    staff_df = staff_df.copy()
    staff_df["Name"] = staff_df[name_c].astype(str).str.strip()
    staff_df["HomeSite"] = staff_df[home_c].astype(str).str.strip().str.upper() if home_c else ""

    # Convert skill flags: accept Y/Yes/True/1
    def yn(v):
        if pd.isna(v): return False
        s = str(v).strip().lower()
        return s in ["y", "yes", "true", "1"]

    # Preserve all Can* columns as booleans
    for c in staff_df.columns:
        if normalize(c).startswith("can") or normalize(c) in ["iscarolchurchyn", "iscarolchurch"]:
            staff_df[c] = staff_df[c].apply(yn)

    # Carol
    carol_c = pick_col(staff_df, ["IsCarolChurch(Y/N)", "IsCarolChurch", "Carol"], required=False)
    if carol_c:
        staff_df["IsCarolChurch"] = staff_df[carol_c].apply(bool)
    else:
        staff_df["IsCarolChurch"] = staff_df["Name"].str.lower().eq("carol church")

    # Hours
    hours_df = hours_df.copy()
    hours_name_c = pick_col(hours_df, ["Name", "StaffName"])
    hours_df["Name"] = hours_df[hours_name_c].astype(str).str.strip()

    for d in ["Mon","Tue","Wed","Thu","Fri"]:
        sc = pick_col(hours_df, [f"{d}Start", f"{d} Start", f"{d}_Start"], required=False)
        ec = pick_col(hours_df, [f"{d}End", f"{d} End", f"{d}_End"], required=False)
        hours_df[f"{d}Start"] = hours_df[sc].apply(to_time) if sc else None
        hours_df[f"{d}End"] = hours_df[ec].apply(to_time) if ec else None

    # Holidays
    hols = []
    if not hols_df.empty:
        hn = pick_col(hols_df, ["Name", "StaffName"], required=False) or hols_df.columns[0]
        hs = pick_col(hols_df, ["StartDate", "Start"], required=False) or hols_df.columns[1]
        he = pick_col(hols_df, ["EndDate", "End"], required=False) or hols_df.columns[2]
        for _, r in hols_df.iterrows():
            hols.append((str(r[hn]).strip(), to_date(r[hs]), to_date(r[he])))

    return staff_df, hours_df, hols, xls.sheet_names


# =========================================================
# Business rules (your latest spec)
# =========================================================
DAY_START = time(8, 0)
DAY_END = time(18, 30)
SLOT_MIN = 30
MIN_STINT_SLOTS = 3  # 1.5 hours
BREAK_WINDOW = (time(12, 0), time(14, 0))
BREAK_LEN_SLOTS = 1  # 30 mins
BREAK_THRESHOLD_HOURS = 6.0

# Mandatory coverage
MANDATORY_RULES = [
    ("FrontDesk_SLGP", DAY_START, DAY_END, 1),
    ("FrontDesk_JEN",  DAY_START, DAY_END, 1),
    ("FrontDesk_BGS",  DAY_START, DAY_END, 1),

    ("Triage_Admin_SLGP", DAY_START, time(16, 0), 1),
    ("Triage_Admin_JEN",  DAY_START, time(16, 0), 1),

    ("Email_Box", DAY_START, DAY_END, 1),

    ("Phones", DAY_START, DAY_END, 2),     # min 2 phones
    ("Bookings", DAY_START, DAY_END, 3),   # min 3 bookings (SLGP)
]

def awaiting_site_for_day(d: date) -> str:
    # Mon/Fri SLGP, Tue/Thu JEN, Wed BGS
    wd = d.weekday()
    if wd in (0, 4):
        return "SLGP"
    if wd in (1, 3):
        return "JEN"
    return "BGS"

def awaiting_required(d: date, t: time) -> bool:
    return t_in_range(t, time(10, 0), time(16, 0))

def awaiting_optional(d: date, t: time) -> bool:
    return t_in_range(t, time(16, 0), DAY_END)

# Filler order (after mandatory)
def filler_order_for_day(d: date):
    # Mon/Fri prioritise Phones over Emis/Docman
    if d.weekday() in (0, 4):
        return ["Phones", "Bookings", "Awaiting_PSA_Admin", "Emis_Tasks", "Docman_Tasks"]
    return ["Phones", "Bookings", "Emis_Tasks", "Docman_Tasks", "Awaiting_PSA_Admin"]


# =========================================================
# Skills mapping to YOUR template columns
# =========================================================
def can(sr, *colnames) -> bool:
    for c in colnames:
        if c in sr.index:
            return bool(sr[c])
    return False

def skill_allowed(sr, role: str) -> bool:
    # Your Staff sheet uses CanFrontDesk / CanTriage / CanEmail / CanPhones / CanBookings / CanEMIS / CanDocman_PSA / CanDocman_AWAIT
    if role.startswith("FrontDesk"):
        return can(sr, "CanFrontDesk")
    if role.startswith("Triage_Admin"):
        return can(sr, "CanTriage")
    if role == "Email_Box":
        return can(sr, "CanEmail")
    if role == "Phones":
        return can(sr, "CanPhones")
    if role == "Bookings":
        return can(sr, "CanBookings")
    if role == "Emis_Tasks":
        return can(sr, "CanEMIS")
    if role == "Docman_Tasks":
        return can(sr, "CanDocman_PSA") or can(sr, "CanDocman_AWAIT")
    if role == "Awaiting_PSA_Admin":
        # Treated as docman/awaiting capability
        return can(sr, "CanDocman_PSA") or can(sr, "CanDocman_AWAIT")
    return False

def site_restriction_ok(sr, role: str, d: date) -> bool:
    home = str(sr.get("HomeSite", "")).strip().upper()

    # Email/Phones/EMIS/Docman: only JEN + BGS (per your original rule)
    if role in ["Email_Box", "Phones", "Emis_Tasks", "Docman_Tasks"]:
        return home in ["JEN", "BGS"]

    # Awaiting/PSA: MUST be on that day's site (Mon/Fri SLGP, Tue/Thu JEN, Wed BGS)
    if role == "Awaiting_PSA_Admin":
        return home == awaiting_site_for_day(d)

    # Bookings: SLGP
    if role == "Bookings":
        return home == "SLGP"

    # Site-specific roles
    if role.endswith("_SLGP"):
        return home == "SLGP"
    if role.endswith("_JEN"):
        return home == "JEN"
    if role.endswith("_BGS"):
        return home == "BGS"

    return True


# =========================================================
# Availability + breaks
# =========================================================
def is_on_holiday(name: str, d: date, hols):
    for n, s, e in hols:
        if n.strip().lower() == name.strip().lower() and s and e and s <= d <= e:
            return True
    return False

def build_availability(staff_df, hours_df, hols, week_start: date):
    hmap = {r["Name"]: r for _, r in hours_df.iterrows()}
    days = ["Mon","Tue","Wed","Thu","Fri"]
    avail = {}
    for i, dname in enumerate(days):
        d = week_start + timedelta(days=i)
        names = []
        for _, s in staff_df.iterrows():
            nm = s["Name"]
            if is_on_holiday(nm, d, hols):
                continue
            hr = hmap.get(nm)
            if hr is None:
                continue
            stt = hr.get(f"{dname}Start")
            end = hr.get(f"{dname}End")
            if stt and end:
                names.append(nm)
        avail[d] = names
    return avail, hmap

def timeslots():
    cur = datetime(2000,1,1,DAY_START.hour,DAY_START.minute)
    end = datetime(2000,1,1,DAY_END.hour,DAY_END.minute)
    out = []
    while cur < end:
        out.append(cur.time())
        cur += timedelta(minutes=SLOT_MIN)
    return out

def shift_window(hrow, dname):
    return hrow.get(f"{dname}Start"), hrow.get(f"{dname}End")

def pick_break_slot_near_midpoint(d: date, stt: time, end: time):
    if not stt or not end:
        return None
    if h_between(stt, end) <= BREAK_THRESHOLD_HOURS:
        return None

    midpoint = dt_of(d, stt) + (dt_of(d, end) - dt_of(d, stt)) / 2
    candidates = []
    for t in [time(12,0), time(12,30), time(13,0), time(13,30)]:
        if t >= stt and add_minutes(t, 30) <= end and t_in_range(t, BREAK_WINDOW[0], BREAK_WINDOW[1]):
            dist = abs((dt_of(d, t) - midpoint).total_seconds())
            candidates.append((dist, t))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0])
    return candidates[0][1]


# =========================================================
# Scheduler (min 1.5h stints)
# =========================================================
def rota_generate_week(staff_df, hours_df, hols, week_start: date):
    slots = timeslots()
    days = ["Mon","Tue","Wed","Thu","Fri"]
    dates = [week_start + timedelta(days=i) for i in range(5)]

    avail, hmap = build_availability(staff_df, hours_df, hols, week_start)
    staff_by_name = {r["Name"]: r for _, r in staff_df.iterrows()}

    # breaks: (d,t)->set(names)
    breaks = {}
    for d in dates:
        dname = days[(d-week_start).days]
        for name in avail.get(d, []):
            hr = hmap[name]
            stt, end = shift_window(hr, dname)
            b = pick_break_slot_near_midpoint(d, stt, end)
            if b:
                breaks.setdefault((d, b), set()).add(name)

    assign = {}  # (d,t)-> list[(role,name)]
    gaps = []

    # Locks to keep minimum stint
    lock = {}  # (d,name)->(role, remaining_slots)

    # Continuous limits
    cont_fd = {}       # (d,name)->hours
    cont_triage = {}   # (d,name)->hours

    # Simple fairness within week
    role_hours = {}    # (name,role)->hours

    def is_working(name, d, t):
        dname = days[(d-week_start).days]
        hr = hmap.get(name)
        if hr is None:
            return False
        stt, end = shift_window(hr, dname)
        return bool(stt and end and (t >= stt) and (t < end))

    def on_break(name, d, t):
        return name in breaks.get((d, t), set())

    def used(name, d, t):
        return any(nm == name for _, nm in assign.get((d, t), []))

    def score(name, role, d):
        base = role_hours.get((name, role), 0.0)
        # prefer staying on same task
        cur = lock.get((d, name))
        if cur and cur[0] == role and cur[1] > 0:
            base -= 0.75
        # deprioritise extra FD load
        if role.startswith("FrontDesk"):
            base += 2.0 * sum(v for (nm, rl), v in role_hours.items() if nm == name and rl.startswith("FrontDesk"))
        return base

    def candidate_ok(name, role, d, t):
        if not is_working(name, d, t):
            return False
        if on_break(name, d, t):
            return False
        if used(name, d, t):
            return False

        sr = staff_by_name[name]

        if not skill_allowed(sr, role):
            return False
        if not site_restriction_ok(sr, role, d):
            return False

        cur = lock.get((d, name))
        if cur and cur[1] > 0 and cur[0] != role:
            return False

        if role.startswith("FrontDesk") and cont_fd.get((d, name), 0.0) + 0.5 > 2.5:
            return False
        if role.startswith("Triage_Admin") and cont_triage.get((d, name), 0.0) + 0.5 > 3.0:
            return False

        return True

    def force_carol_if_needed(role, d, t):
        if not role.startswith("FrontDesk"):
            return None
        for _, sr in staff_df.iterrows():
            if sr.get("IsCarolChurch", False):
                nm = sr["Name"]
                if is_working(nm, d, t) and not on_break(nm, d, t):
                    home = str(sr.get("HomeSite","")).strip().upper()
                    if role.endswith(home):
                        return nm
        return None

    for d in dates:
        for t in slots:
            assign.setdefault((d, t), [])

            # Mandatory roles for this slot
            wanted = [(role, need) for role, a, b, need in MANDATORY_RULES if t_in_range(t, a, b)]
            # Awaiting mandatory 10-16 and optional after 16 if possible
            if awaiting_required(d, t) or awaiting_optional(d, t):
                wanted.append(("Awaiting_PSA_Admin", 1))

            # Fill mandatory
            for role, need in wanted:
                already = sum(1 for rl, _ in assign[(d, t)] if rl == role)
                while already < need:
                    forced = force_carol_if_needed(role, d, t)
                    if forced and candidate_ok(forced, role, d, t):
                        pick = forced
                    else:
                        cands = [n for n in avail.get(d, []) if candidate_ok(n, role, d, t)]
                        if not cands:
                            # Only mark as gap if role is truly mandatory at this time
                            if role != "Awaiting_PSA_Admin" or awaiting_required(d, t):
                                gaps.append((d, t, role, "No suitable staff available"))
                            break
                        cands.sort(key=lambda n: score(n, role, d))
                        pick = cands[0]

                    assign[(d, t)].append((role, pick))

                    # lock for minimum stint
                    cur = lock.get((d, pick))
                    if (not cur) or (cur[0] != role) or (cur[1] <= 0):
                        lock[(d, pick)] = (role, MIN_STINT_SLOTS - 1)

                    # decrement lock for those currently assigned in this slot happens next slot implicitly
                    # update continuous counters
                    if role.startswith("FrontDesk"):
                        cont_fd[(d, pick)] = cont_fd.get((d, pick), 0.0) + 0.5
                    else:
                        cont_fd[(d, pick)] = 0.0
                    if role.startswith("Triage_Admin"):
                        cont_triage[(d, pick)] = cont_triage.get((d, pick), 0.0) + 0.5
                    else:
                        cont_triage[(d, pick)] = 0.0

                    role_hours[(pick, role)] = role_hours.get((pick, role), 0.0) + 0.5
                    already += 1

            # Decrement locks for staff that are assigned and locked
            for role, name in assign[(d, t)]:
                cur = lock.get((d, name))
                if cur and cur[1] > 0 and cur[0] == role:
                    lock[(d, name)] = (cur[0], cur[1] - 1)

            # Fillers: assign idle staff (not mandatory) in priority order
            filler_order = filler_order_for_day(d)
            max_phones = 3
            current_phones = sum(1 for rl, _ in assign[(d, t)] if rl == "Phones")

            for role in filler_order:
                if role == "Phones" and current_phones >= max_phones:
                    continue

                # optional Awaiting after 16:00 only (mandatory already handled above)
                if role == "Awaiting_PSA_Admin" and not awaiting_optional(d, t):
                    continue

                # determine how many to add
                target = 0
                if role == "Phones":
                    target = max_phones - current_phones
                else:
                    target = 999

                for _ in range(target):
                    cands = []
                    for name in avail.get(d, []):
                        if used(name, d, t) or on_break(name, d, t) or not is_working(name, d, t):
                            continue
                        cur = lock.get((d, name))
                        if cur and cur[1] > 0 and cur[0] != role:
                            continue
                        sr = staff_by_name[name]
                        if not skill_allowed(sr, role):
                            continue
                        if not site_restriction_ok(sr, role, d):
                            continue
                        cands.append(name)

                    if not cands:
                        break

                    # Enforce "majority of SLGP on bookings": if SLGP and trained for bookings, keep them on Bookings as filler
                    if role != "Bookings":
                        filtered = []
                        for n in cands:
                            sr = staff_by_name[n]
                            if str(sr.get("HomeSite","")).upper() == "SLGP" and skill_allowed(sr, "Bookings"):
                                continue
                            filtered.append(n)
                        if filtered:
                            cands = filtered

                    cands.sort(key=lambda n: score(n, role, d))
                    pick = cands[0]

                    assign[(d, t)].append((role, pick))
                    cur = lock.get((d, pick))
                    if (not cur) or (cur[0] != role) or (cur[1] <= 0):
                        lock[(d, pick)] = (role, MIN_STINT_SLOTS - 1)

                    role_hours[(pick, role)] = role_hours.get((pick, role), 0.0) + 0.5
                    if role == "Phones":
                        current_phones += 1

    # Break check
    for d in dates:
        dname = days[(d-week_start).days]
        for name in avail.get(d, []):
            hr = hmap[name]
            stt, end = shift_window(hr, dname)
            if not stt or not end:
                continue
            if h_between(stt, end) <= BREAK_THRESHOLD_HOURS:
                continue
            had_break = any(name in breaks.get((d, t), set()) for t in [time(12,0), time(12,30), time(13,0), time(13,30)])
            if not had_break:
                gaps.append((d, None, "Break", f"{name}: shift > 6h but no break could be placed 12:00–14:00"))

    return assign, breaks, gaps


# =========================================================
# Excel output (Grid + Gaps + Totals + Timelines by site)
# =========================================================
ROLE_COLORS = {
    "FrontDesk_SLGP": "FFF2CC",
    "FrontDesk_JEN":  "FFF2CC",
    "FrontDesk_BGS":  "FFF2CC",
    "Triage_Admin_SLGP": "D9EAD3",
    "Triage_Admin_JEN":  "D9EAD3",
    "Email_Box": "CFE2F3",
    "Phones": "C9DAF8",
    "Bookings": "FCE5CD",
    "Emis_Tasks": "EAD1DC",
    "Docman_Tasks": "D0E0E3",
    "Awaiting_PSA_Admin": "D0E0E3",
    "Break": "E6E6E6",
}

def fill_for_role(role):
    return PatternFill("solid", fgColor=ROLE_COLORS.get(role, "FFFFFF"))

def build_timeline_map(assign, breaks, week_start: date, staff_names):
    slots = timeslots()
    dates = [week_start + timedelta(days=i) for i in range(5)]
    m = {}
    for d in dates:
        for t in slots:
            for nm in staff_names:
                if nm in breaks.get((d, t), set()):
                    m[(d, t, nm)] = "Break"
                else:
                    roles = [rl for rl, n in assign.get((d, t), []) if n == nm]
                    m[(d, t, nm)] = " + ".join(roles) if roles else ""
    return m

def write_week(wb: Workbook, week_start: date, staff_df, assign, breaks, gaps, w_idx: int):
    slots = timeslots()
    dates = [week_start + timedelta(days=i) for i in range(5)]
    staff_names = list(staff_df["Name"].astype(str))

    # Grid
    ws_grid = wb.create_sheet(f"Week{w_idx}_Grid")
    ws_grid.append(["Time"] + [d.strftime("%a %d-%b") for d in dates])
    for c in ws_grid[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

    for t in slots:
        row = [t.strftime("%H:%M")]
        for d in dates:
            slot_roles = assign.get((d, t), [])
            parts = []
            for role in [
                "FrontDesk_SLGP","FrontDesk_JEN","FrontDesk_BGS",
                "Triage_Admin_SLGP","Triage_Admin_JEN",
                "Email_Box","Phones","Bookings",
                "Awaiting_PSA_Admin","Emis_Tasks","Docman_Tasks",
            ]:
                ppl = [nm for rl, nm in slot_roles if rl == role]
                if ppl:
                    parts.append(f"{role}: " + ", ".join(ppl))
            row.append("\n".join(parts))
        ws_grid.append(row)

    ws_grid.column_dimensions["A"].width = 8
    for col in range(2, 7):
        ws_grid.column_dimensions[chr(64+col)].width = 42
    for r in range(2, 2+len(slots)):
        ws_grid.row_dimensions[r].height = 95
        for c in range(2, 7):
            ws_grid.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")

    # Coverage gaps
    ws_gaps = wb.create_sheet(f"Week{w_idx}_CoverageGaps")
    ws_gaps.append(["Date", "Time", "Role", "Issue"])
    for c in ws_gaps[1]:
        c.font = Font(bold=True)
    for d, t, role, issue in gaps:
        ws_gaps.append([d.isoformat(), "" if t is None else t.strftime("%H:%M"), role, issue])

    # Totals
    ws_tot = wb.create_sheet(f"Week{w_idx}_Totals")
    rows = []
    for d in dates:
        for t in slots:
            for role, nm in assign.get((d, t), []):
                rows.append([d, nm, role, 0.5])
            for nm in breaks.get((d, t), set()):
                rows.append([d, nm, "Break", 0.5])

    df = pd.DataFrame(rows, columns=["Date", "Name", "Task", "Hours"])
    if df.empty:
        df = pd.DataFrame(columns=["Date","Name","Task","Hours"])

    pivot_w = (df.groupby(["Name", "Task"])["Hours"].sum()
                 .reset_index()
                 .pivot(index="Name", columns="Task", values="Hours")
                 .fillna(0.0))
    pivot_w["WeeklyTotal"] = pivot_w.sum(axis=1)
    pivot_w = pivot_w.reset_index()

    pivot_d = df.groupby(["Date", "Name", "Task"])["Hours"].sum().reset_index()

    ws_tot.append(["Weekly totals per staff by task (hours)"])
    ws_tot["A1"].font = Font(bold=True, size=12)
    ws_tot.append([])
    for r in dataframe_to_rows(pivot_w, index=False, header=True):
        ws_tot.append(r)

    header_row = 3
    headers = [c.value for c in ws_tot[header_row]]
    for cell in ws_tot[header_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for j, h in enumerate(headers, start=1):
        if not h or h == "Name":
            continue
        fill = fill_for_role(h)
        for irow in range(header_row, header_row + len(pivot_w) + 1):
            ws_tot.cell(irow, j).fill = fill

    ws_tot.append([])
    ws_tot.append(["Daily totals per staff by task (hours)"])
    ws_tot[f"A{ws_tot.max_row}"].font = Font(bold=True, size=12)
    ws_tot.append([])
    for r in dataframe_to_rows(pivot_d, index=False, header=True):
        ws_tot.append(r)

    # Timelines
    m = build_timeline_map(assign, breaks, week_start, staff_names)

    ws_all = wb.create_sheet(f"Week{w_idx}_Timelines_All")
    ws_all.append(["Date", "Time"] + staff_names)
    for c in ws_all[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws_all.freeze_panes = "C2"

    for d in dates:
        for t in slots:
            row = [d.strftime("%a %d-%b"), t.strftime("%H:%M")]
            for nm in staff_names:
                row.append(m.get((d, t, nm), ""))
            ws_all.append(row)

    for r in range(2, ws_all.max_row + 1):
        for c in range(3, ws_all.max_column + 1):
            val = ws_all.cell(r, c).value
            if val:
                role = val.split(" + ")[0]
                ws_all.cell(r, c).fill = fill_for_role(role)
                ws_all.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")

    # By site tabs
    for site in ["SLGP", "JEN", "BGS"]:
        site_staff = list(staff_df.loc[staff_df["HomeSite"].astype(str).str.upper() == site, "Name"].astype(str))
        if not site_staff:
            continue
        ws_site = wb.create_sheet(f"Week{w_idx}_{site}_Timelines")
        ws_site.append(["Date", "Time"] + site_staff)
        for c in ws_site[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_site.freeze_panes = "C2"

        for d in dates:
            for t in slots:
                row = [d.strftime("%a %d-%b"), t.strftime("%H:%M")]
                for nm in site_staff:
                    row.append(m.get((d, t, nm), ""))
                ws_site.append(row)

        for r in range(2, ws_site.max_row + 1):
            for c in range(3, ws_site.max_column + 1):
                val = ws_site.cell(r, c).value
                if val:
                    role = val.split(" + ")[0]
                    ws_site.cell(r, c).fill = fill_for_role(role)
                    ws_site.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")


def build_workbook(staff_df, hours_df, hols, start_monday: date, weeks: int):
    wb = Workbook()
    wb.remove(wb.active)

    for w in range(weeks):
        ws = start_monday + timedelta(days=7*w)
        assign, breaks, gaps = rota_generate_week(staff_df, hours_df, hols, ws)
        write_week(wb, ws, staff_df, assign, breaks, gaps, w_idx=w+1)

    return wb


# =========================================================
# Streamlit UI
# =========================================================
st.set_page_config(page_title="Rota Generator", layout="wide")
require_password()

st.title("Rota Generator (Excel export)")

uploaded = st.file_uploader("Upload rota template (.xlsx)", type=["xlsx"])

c1, c2 = st.columns(2)
with c1:
    start_date = st.date_input("Week commencing (Monday)", value=date.today())
with c2:
    mode = st.selectbox("Run period", ["1 week", "4 weeks (month)", "Custom # weeks"])

if mode == "Custom # weeks":
    weeks = int(st.number_input("Number of weeks", min_value=1, max_value=12, value=2, step=1))
elif mode == "4 weeks (month)":
    weeks = 4
else:
    weeks = 1

start_monday = start_date - timedelta(days=start_date.weekday())

st.caption(
    "Key rules: minimum 1.5h stints; breaks between 12:00–14:00 chosen closest to shift midpoint; "
    "Front Desk max 2.5h continuous; Triage max 3h continuous; "
    "Phones min 2 (JEN/BGS only); Bookings min 3 (SLGP only, and SLGP filler time is pushed to Bookings); "
    "Awaiting Response/PSA Admin is mandatory 10:00–16:00 on: Mon/Fri SLGP, Tue/Thu JEN, Wed BGS (home site must match). "
    "Docman PSA and Awaiting are treated as Docman Tasks."
)

if uploaded:
    try:
        staff_df, hours_df, hols, sheets = read_template(uploaded.getvalue())
        st.success("Template loaded.")

        if st.button("Generate rota and download Excel", type="primary"):
            wb = build_workbook(staff_df, hours_df, hols, start_monday, weeks)
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            out_name = f"rota_{start_monday.isoformat()}_{weeks}w.xlsx"
            st.download_button(
                "📊 Download Excel rota",
                data=bio.getvalue(),
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    except Exception as e:
        st.error("Could not process the template.")
        st.exception(e)
else:
    st.info("Upload your completed template to continue.")
