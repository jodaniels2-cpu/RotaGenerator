import io
import re
from datetime import datetime, date, time, timedelta

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================================================
# Password (Streamlit secrets TOML)
# =========================================================
def require_password():
    """
    In Streamlit Secrets (TOML):
      APP_PASSWORD = "your-strong-password"
    """
    pw = st.secrets.get("APP_PASSWORD")
    if not pw:
        st.warning("APP_PASSWORD not set in Streamlit Secrets. App is not password-protected.")
        return True
    if st.session_state.get("authed"):
        return True
    with st.form("login"):
        entered = st.text_input("Password", type="password")
        ok = st.form_submit_button("Log in")
        if ok and entered == pw:
            st.session_state.authed = True
            st.success("Logged in.")
            return True
        if ok:
            st.error("Incorrect password.")
    st.stop()

# =========================================================
# Helpers
# =========================================================
def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s).strip().lower())

def find_sheet(xls: pd.ExcelFile, candidates):
    names = {normalize(n): n for n in xls.sheet_names}
    for c in candidates:
        key = normalize(c)
        if key in names:
            return names[key]
    for n in xls.sheet_names:
        nn = normalize(n)
        for c in candidates:
            if normalize(c) in nn or nn in normalize(c):
                return n
    return None

def pick_col(df: pd.DataFrame, candidates, required=True):
    cols = {normalize(c): c for c in df.columns}
    for cand in candidates:
        key = normalize(cand)
        if key in cols:
            return cols[key]
    for c in df.columns:
        nc = normalize(c)
        for cand in candidates:
            if normalize(cand) in nc:
                return c
    if required:
        raise KeyError(f"Missing required column among {candidates}. Available: {list(df.columns)}")
    return None

def to_time(x):
    if pd.isna(x) or x == "":
        return None
    if isinstance(x, time):
        return x
    if isinstance(x, datetime):
        return x.time()
    if isinstance(x, (float, int)):
        seconds = int(round(float(x) * 86400))
        return (datetime(2000, 1, 1) + timedelta(seconds=seconds)).time()
    return pd.to_datetime(str(x)).time()

def to_date(x):
    if pd.isna(x) or x == "":
        return None
    if isinstance(x, (datetime, pd.Timestamp)):
        return x.date()
    if isinstance(x, date):
        return x
    return pd.to_datetime(x).date()

def dt_of(d: date, t: time) -> datetime:
    return datetime(d.year, d.month, d.day, t.hour, t.minute)

def add_minutes(t: time, mins: int) -> time:
    return (datetime(2000, 1, 1, t.hour, t.minute) + timedelta(minutes=mins)).time()

def t_in_range(t: time, a: time, b: time) -> bool:
    return (t >= a) and (t < b)

def h_between(t1: time, t2: time) -> float:
    return (dt_of(date(2000,1,1), t2) - dt_of(date(2000,1,1), t1)).total_seconds() / 3600.0

# =========================================================
# Read template (robust + holiday notes)
# =========================================================
def read_template(uploaded_bytes: bytes):
    xls = pd.ExcelFile(io.BytesIO(uploaded_bytes))
    staff_sheet = find_sheet(xls, ["Staff", "Skills", "People"])
    hours_sheet = find_sheet(xls, ["WorkingHours", "Hours", "Availability"])
    hols_sheet = find_sheet(xls, ["Holidays", "Leave", "Absence"])

    if not staff_sheet:
        raise ValueError(f"Could not find Staff sheet. Found: {xls.sheet_names}")
    if not hours_sheet:
        raise ValueError(f"Could not find WorkingHours sheet. Found: {xls.sheet_names}")

    staff_df = pd.read_excel(xls, sheet_name=staff_sheet)
    hours_df = pd.read_excel(xls, sheet_name=hours_sheet)
    hols_df = pd.read_excel(xls, sheet_name=hols_sheet) if hols_sheet else pd.DataFrame()

    name_c = pick_col(staff_df, ["Name", "StaffName"])
    home_c = pick_col(staff_df, ["HomeSite", "Site", "BaseSite"], required=False)

    staff_df = staff_df.copy()
    staff_df["Name"] = staff_df[name_c].astype(str).str.strip()
    staff_df["HomeSite"] = staff_df[home_c].astype(str).str.strip().str.upper() if home_c else ""

    def yn(v):
        if pd.isna(v): return False
        s = str(v).strip().lower()
        return s in ["y", "yes", "true", "1"]

    for c in staff_df.columns:
        if normalize(c).startswith("can") or normalize(c) in ["iscarolchurchyn", "iscarolchurch"]:
            staff_df[c] = staff_df[c].apply(yn)

    carol_c = pick_col(staff_df, ["IsCarolChurch(Y/N)", "IsCarolChurch", "Carol"], required=False)
    if carol_c:
        staff_df["IsCarolChurch"] = staff_df[carol_c].apply(bool)
    else:
        staff_df["IsCarolChurch"] = staff_df["Name"].str.lower().eq("carol church")

    hours_df = hours_df.copy()
    hours_name_c = pick_col(hours_df, ["Name", "StaffName"])
    hours_df["Name"] = hours_df[hours_name_c].astype(str).str.strip()

    for d in ["Mon","Tue","Wed","Thu","Fri"]:
        sc = pick_col(hours_df, [f"{d}Start", f"{d} Start", f"{d}_Start"], required=False)
        ec = pick_col(hours_df, [f"{d}End", f"{d} End", f"{d}_End"], required=False)
        hours_df[f"{d}Start"] = hours_df[sc].apply(to_time) if sc else None
        hours_df[f"{d}End"] = hours_df[ec].apply(to_time) if ec else None

    # Holidays with Notes => Holiday default unless "sickness" or "bank holiday"
    hols = []
    if not hols_df.empty:
        hn = pick_col(hols_df, ["Name", "StaffName"], required=False) or hols_df.columns[0]
        hs = pick_col(hols_df, ["StartDate", "Start"], required=False) or hols_df.columns[1]
        he = pick_col(hols_df, ["EndDate", "End"], required=False) or hols_df.columns[2]
        notes_c = pick_col(hols_df, ["Notes", "Note", "Reason"], required=False)

        for _, r in hols_df.iterrows():
            nm = str(r[hn]).strip()
            sd = to_date(r[hs])
            ed = to_date(r[he])
            note = "" if (not notes_c or pd.isna(r[notes_c])) else str(r[notes_c]).strip().lower()
            kind = "Holiday"
            if "sick" in note or "sickness" in note:
                kind = "Sick"
            elif "bank" in note:
                kind = "Bank Holiday"
            hols.append((nm, sd, ed, kind))

    return staff_df, hours_df, hols

# =========================================================
# Rules
# =========================================================
DAY_START = time(8, 0)
DAY_END = time(18, 30)
SLOT_MIN = 30

MIN_STINT_SLOTS = 4   # 2 hours
MAX_STINT_SLOTS = 6   # 3 hours (Email can exceed)
BREAK_WINDOW = (time(12, 0), time(14, 0))
BREAK_THRESHOLD_HOURS = 6.0

MANDATORY_RULES = [
    ("FrontDesk_SLGP", DAY_START, DAY_END, 1),
    ("FrontDesk_JEN",  DAY_START, DAY_END, 1),
    ("FrontDesk_BGS",  DAY_START, DAY_END, 1),

    ("Triage_Admin_SLGP", DAY_START, time(16, 0), 1),
    ("Triage_Admin_JEN",  DAY_START, time(16, 0), 1),

    ("Email_Box", DAY_START, DAY_END, 1),

    ("Phones", DAY_START, DAY_END, 2),
    ("Bookings", DAY_START, DAY_END, 3),
]

def awaiting_site_for_day(d: date) -> str:
    wd = d.weekday()
    if wd in (0, 4):
        return "SLGP"
    if wd in (1, 3):
        return "JEN"
    return "BGS"

def awaiting_required(d: date, t: time) -> bool:
    return t_in_range(t, time(10, 0), time(16, 0))

def awaiting_optional(d: date, t: time) -> bool:
    return t_in_range(t, time(16, 0), DAY_END)

def filler_order_for_day(d: date):
    if d.weekday() in (0, 4):
        return ["Phones", "Bookings", "Awaiting_PSA_Admin", "Emis_Tasks", "Docman_Tasks"]
    return ["Phones", "Bookings", "Emis_Tasks", "Docman_Tasks", "Awaiting_PSA_Admin"]

# =========================================================
# Skills
# =========================================================
def can(sr, *cols) -> bool:
    for c in cols:
        if c in sr.index:
            return bool(sr[c])
    return False

def skill_allowed(sr, role: str) -> bool:
    if role.startswith("FrontDesk"):
        return can(sr, "CanFrontDesk")
    if role.startswith("Triage_Admin"):
        return can(sr, "CanTriage")
    if role == "Email_Box":
        return can(sr, "CanEmail")
    if role == "Phones":
        return can(sr, "CanPhones")
    if role == "Bookings":
        return can(sr, "CanBookings")
    if role == "Emis_Tasks":
        return can(sr, "CanEMIS")
    if role == "Docman_Tasks":
        return can(sr, "CanDocman_PSA") or can(sr, "CanDocman_AWAIT")
    if role == "Awaiting_PSA_Admin":
        return can(sr, "CanDocman_PSA") or can(sr, "CanDocman_AWAIT")
    if role == "Unassigned":
        return True
    return False

def role_site(role: str, d: date) -> str:
    if role.endswith("_SLGP"):
        return "SLGP"
    if role.endswith("_JEN"):
        return "JEN"
    if role.endswith("_BGS"):
        return "BGS"
    if role == "Bookings":
        return "SLGP"
    if role == "Awaiting_PSA_Admin":
        return awaiting_site_for_day(d)
    return ""  # shared

def site_allowed(sr, role: str, d: date, allow_cross: bool) -> bool:
    home = str(sr.get("HomeSite", "")).strip().upper()

    # Never cross-site for front desk
    if role.startswith("FrontDesk"):
        return home == role_site(role, d)

    # Bookings SLGP only
    if role == "Bookings":
        return home == "SLGP"

    # Site-bound triage roles
    if role.startswith("Triage_Admin"):
        target = role_site(role, d)
        return (home == target) or allow_cross

    # Awaiting is site by weekday
    if role == "Awaiting_PSA_Admin":
        target = awaiting_site_for_day(d)
        return (home == target) or allow_cross

    # Email/Phones/Emis/Docman normally JEN/BGS only, but can be covered cross-site if no home-site staff exist
    if role in ["Email_Box", "Phones", "Emis_Tasks", "Docman_Tasks"]:
        if home in ["JEN", "BGS"]:
            return True
        return allow_cross

    return True

# =========================================================
# Availability / breaks / holidays
# =========================================================
def timeslots():
    cur = datetime(2000,1,1,DAY_START.hour,DAY_START.minute)
    end = datetime(2000,1,1,DAY_END.hour,DAY_END.minute)
    out = []
    while cur < end:
        out.append(cur.time())
        cur += timedelta(minutes=SLOT_MIN)
    return out

def build_maps(staff_df, hours_df):
    staff_by_name = {r["Name"]: r for _, r in staff_df.iterrows()}
    hmap = {r["Name"]: r for _, r in hours_df.iterrows()}
    return staff_by_name, hmap

def holiday_kind(name: str, d: date, hols):
    for n, s, e, k in hols:
        if n.strip().lower() == name.strip().lower() and s and e and s <= d <= e:
            return k
    return None

def build_availability(staff_df, hours_df, hols, week_start: date):
    staff_by_name, hmap = build_maps(staff_df, hours_df)
    days = ["Mon","Tue","Wed","Thu","Fri"]
    avail = {}
    for i, dname in enumerate(days):
        d = week_start + timedelta(days=i)
        avail[d] = []
        for name, sr in staff_by_name.items():
            if holiday_kind(name, d, hols):
                continue
            hr = hmap.get(name)
            if hr is None:
                continue
            stt = hr.get(f"{dname}Start")
            end = hr.get(f"{dname}End")
            if stt and end:
                avail[d].append(name)
    return avail, hmap

def is_working(hmap, week_start, d, t, name):
    days = ["Mon","Tue","Wed","Thu","Fri"]
    dname = days[(d-week_start).days]
    hr = hmap.get(name)
    if hr is None:
        return False
    stt = hr.get(f"{dname}Start")
    end = hr.get(f"{dname}End")
    return bool(stt and end and (t >= stt) and (t < end))

def pick_break_slot_near_midpoint(d: date, stt: time, end: time):
    if not stt or not end:
        return None
    if h_between(stt, end) <= BREAK_THRESHOLD_HOURS:
        return None
    midpoint = dt_of(d, stt) + (dt_of(d, end) - dt_of(d, stt)) / 2
    candidates = []
    for t in [time(12,0), time(12,30), time(13,0), time(13,30)]:
        if t >= stt and add_minutes(t, 30) <= end and t_in_range(t, BREAK_WINDOW[0], BREAK_WINDOW[1]):
            dist = abs((dt_of(d, t) - midpoint).total_seconds())
            candidates.append((dist, t))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0])
    return candidates[0][1]

# =========================================================
# Scheduler
# =========================================================
def rota_generate_week(staff_df, hours_df, hols, week_start: date):
    slots = timeslots()
    days = ["Mon","Tue","Wed","Thu","Fri"]
    dates = [week_start + timedelta(days=i) for i in range(5)]
    avail, hmap = build_availability(staff_df, hours_df, hols, week_start)
    staff_by_name, _ = build_maps(staff_df, hours_df)

    # breaks
    breaks = {}
    for d in dates:
        dname = days[(d-week_start).days]
        for name in avail.get(d, []):
            hr = hmap[name]
            stt = hr.get(f"{dname}Start")
            end = hr.get(f"{dname}End")
            b = pick_break_slot_near_midpoint(d, stt, end)
            if b:
                breaks.setdefault((d, b), set()).add(name)

    assign = {}  # (d,t)->list[(role,name)]
    gaps = []

    # stint tracking per day (site sticky via last_site)
    cur_role = {}
    stint = {}
    lock_rem = {}
    last_site = {}  # site of last assignment (for site stickiness)

    role_hours = {}

    def on_break(name, d, t): return name in breaks.get((d, t), set())
    def used(name, d, t): return any(nm == name for _, nm in assign.get((d, t), []))

    def must_switch(name, d):
        role = cur_role.get((d, name))
        if role == "Email_Box":
            return False
        return stint.get((d, name), 0) >= MAX_STINT_SLOTS

    def can_continue(name, d, role):
        if role == "Email_Box":
            return True
        return not must_switch(name, d)

    def site_sticky_penalty(name, d, role):
        # Penalize changing site within a day (FrontDesk/Triage/Awaiting/Bookings are site-bound)
        s = role_site(role, d)
        prev = last_site.get((d, name))
        if not s:
            return 0.0
        if prev and prev != s:
            return 5.0
        return 0.0

    def score(name, role, d):
        base = role_hours.get((name, role), 0.0)
        if cur_role.get((d, name)) == role and can_continue(name, d, role):
            base -= 1.5
        base += site_sticky_penalty(name, d, role)
        if role.startswith("FrontDesk"):
            base += 2.0 * sum(v for (nm, rl), v in role_hours.items() if nm == name and rl.startswith("FrontDesk"))
        return base

    def candidate_ok(name, role, d, t, allow_cross):
        if not is_working(hmap, week_start, d, t, name):
            return False
        if on_break(name, d, t):
            return False
        if used(name, d, t):
            return False

        sr = staff_by_name[name]
        if not skill_allowed(sr, role):
            return False

        # site stickiness: if locked to a site already, avoid switching unless cross is essential and role has site
        if role_site(role, d):
            prev = last_site.get((d, name))
            if prev and prev != role_site(role, d):
                # allow only if cross-site is enabled AND they are not locked by min-stint
                if lock_rem.get((d, name), 0) > 0:
                    return False

        if not site_allowed(sr, role, d, allow_cross=allow_cross):
            return False

        # min-stint lock
        prev_role = cur_role.get((d, name))
        if prev_role and prev_role != role and lock_rem.get((d, name), 0) > 0:
            return False

        # max-stint
        if prev_role == role and must_switch(name, d):
            return False

        return True

    def pick_candidate(role, d, t):
        # home-only first
        home_cands = [n for n in avail.get(d, []) if candidate_ok(n, role, d, t, allow_cross=False)]
        if home_cands:
            home_cands.sort(key=lambda n: score(n, role, d))
            return home_cands[0]
        # cross-site only if nobody home-site can do it
        cross_cands = [n for n in avail.get(d, []) if candidate_ok(n, role, d, t, allow_cross=True)]
        if cross_cands:
            cross_cands.sort(key=lambda n: score(n, role, d))
            return cross_cands[0]
        return None

    def force_carol(role, d, t):
        if not role.startswith("FrontDesk"):
            return None
        for name, sr in staff_by_name.items():
            if sr.get("IsCarolChurch", False):
                if is_working(hmap, week_start, d, t, name) and not on_break(name, d, t):
                    if role.endswith(sr.get("HomeSite","")):
                        return name
        return None

    for d in dates:
        for t in slots:
            assign.setdefault((d, t), [])

            # Mandatory roles
            wanted = [(role, need) for role, a, b, need in MANDATORY_RULES if t_in_range(t, a, b)]
            if awaiting_required(d, t) or awaiting_optional(d, t):
                wanted.append(("Awaiting_PSA_Admin", 1))

            for role, need in wanted:
                already = sum(1 for rl, _ in assign[(d, t)] if rl == role)
                while already < need:
                    forced = force_carol(role, d, t)
                    if forced and candidate_ok(forced, role, d, t, allow_cross=False):
                        pick = forced
                    else:
                        pick = pick_candidate(role, d, t)
                    if not pick:
                        if role != "Awaiting_PSA_Admin" or awaiting_required(d, t):
                            gaps.append((d, t, role, "No suitable staff available"))
                        break

                    assign[(d, t)].append((role, pick))

                    # update stint & lock
                    prev = cur_role.get((d, pick))
                    if prev == role:
                        stint[(d, pick)] = stint.get((d, pick), 0) + 1
                    else:
                        cur_role[(d, pick)] = role
                        stint[(d, pick)] = 1
                        lock_rem[(d, pick)] = MIN_STINT_SLOTS - 1

                    if lock_rem.get((d, pick), 0) > 0:
                        lock_rem[(d, pick)] -= 1

                    s = role_site(role, d)
                    if s:
                        last_site[(d, pick)] = s

                    role_hours[(pick, role)] = role_hours.get((pick, role), 0.0) + 0.5
                    already += 1

            # Fillers: ensure every working person has a task each slot (or Break)
            filler_roles = filler_order_for_day(d)

            # compute idle
            idle = []
            for name in avail.get(d, []):
                if used(name, d, t) or on_break(name, d, t):
                    continue
                if is_working(hmap, week_start, d, t, name):
                    idle.append(name)

            for name in idle:
                chosen = None
                # If in min-stint, continue role
                if lock_rem.get((d, name), 0) > 0:
                    chosen = cur_role.get((d, name), None)

                # If max-stint hit, force a different role (except Email)
                if chosen and not can_continue(name, d, chosen):
                    chosen = None

                if not chosen:
                    # if SLGP and bookings trained: push bookings
                    sr = staff_by_name[name]
                    if sr.get("HomeSite","") == "SLGP" and skill_allowed(sr, "Bookings"):
                        if candidate_ok(name, "Bookings", d, t, allow_cross=False):
                            chosen = "Bookings"

                if not chosen:
                    for role in filler_roles:
                        if role == "Awaiting_PSA_Admin" and not awaiting_optional(d, t):
                            continue
                        # avoid "same all day" by preferring a different role once min-stint satisfied
                        if cur_role.get((d, name)) == role and stint.get((d, name), 0) >= MAX_STINT_SLOTS and role != "Email_Box":
                            continue
                        # try home then cross if needed
                        if candidate_ok(name, role, d, t, allow_cross=False) or candidate_ok(name, role, d, t, allow_cross=True):
                            chosen = role
                            break

                if not chosen:
                    chosen = "Unassigned"

                assign[(d, t)].append((chosen, name))

                prev = cur_role.get((d, name))
                if prev == chosen:
                    stint[(d, name)] = stint.get((d, name), 0) + 1
                else:
                    cur_role[(d, name)] = chosen
                    stint[(d, name)] = 1
                    lock_rem[(d, name)] = MIN_STINT_SLOTS - 1

                if lock_rem.get((d, name), 0) > 0:
                    lock_rem[(d, name)] -= 1

                s = role_site(chosen, d)
                if s:
                    last_site[(d, name)] = s

                role_hours[(name, chosen)] = role_hours.get((name, chosen), 0.0) + 0.5

    # Break check
    for d in dates:
        dname = days[(d-week_start).days]
        for name in avail.get(d, []):
            hr = hmap[name]
            stt = hr.get(f"{dname}Start")
            end = hr.get(f"{dname}End")
            if not stt or not end:
                continue
            if h_between(stt, end) <= BREAK_THRESHOLD_HOURS:
                continue
            had_break = any(name in breaks.get((d, bt), set()) for bt in [time(12,0), time(12,30), time(13,0), time(13,30)])
            if not had_break:
                gaps.append((d, None, "Break", f"{name}: shift > 6h but no break could be placed 12:00–14:00"))

    return assign, breaks, gaps, hmap

# =========================================================
# Excel output
# =========================================================
ROLE_COLORS = {
    "FrontDesk_SLGP": "FFF2CC",
    "FrontDesk_JEN":  "FFF2CC",
    "FrontDesk_BGS":  "FFF2CC",
    "Triage_Admin_SLGP": "D9EAD3",
    "Triage_Admin_JEN":  "D9EAD3",
    "Email_Box": "CFE2F3",
    "Phones": "C9DAF8",
    "Bookings": "FCE5CD",
    "Emis_Tasks": "EAD1DC",
    "Docman_Tasks": "D0E0E3",
    "Awaiting_PSA_Admin": "D0E0E3",
    "Break": "CFE2F3",
    "Unassigned": "FFFFFF",
    "Not working": "DDDDDD",
    "Holiday": "FFF2CC",
    "Bank Holiday": "FFE599",
    "Sick": "F4CCCC",
}

def fill_for(value: str):
    return PatternFill("solid", fgColor=ROLE_COLORS.get(value, "FFFFFF"))

def build_workbook(staff_df, hours_df, hols, start_monday: date, weeks: int):
    wb = Workbook()
    wb.remove(wb.active)
    slots = timeslots()
    days = ["Mon","Tue","Wed","Thu","Fri"]

    for w in range(weeks):
        week_start = start_monday + timedelta(days=7*w)
        assign, breaks, gaps, hmap = rota_generate_week(staff_df, hours_df, hols, week_start)
        dates = [week_start + timedelta(days=i) for i in range(5)]
        staff_names = list(staff_df["Name"].astype(str))

        # Coverage grid
        ws_grid = wb.create_sheet(f"Week{w+1}_Grid")
        ws_grid.append(["Time"] + [d.strftime("%a %d-%b") for d in dates])
        for c in ws_grid[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")

        for t in slots:
            row = [t.strftime("%H:%M")]
            for d in dates:
                slot_roles = assign.get((d, t), [])
                parts = []
                for role in [
                    "FrontDesk_SLGP","FrontDesk_JEN","FrontDesk_BGS",
                    "Triage_Admin_SLGP","Triage_Admin_JEN",
                    "Email_Box","Phones","Bookings",
                    "Awaiting_PSA_Admin","Emis_Tasks","Docman_Tasks",
                ]:
                    ppl = [nm for rl, nm in slot_roles if rl == role]
                    if ppl:
                        parts.append(f"{role}: " + ", ".join(ppl))
                row.append("\n".join(parts))
            ws_grid.append(row)

        ws_grid.column_dimensions["A"].width = 8
        for col in range(2, 7):
            ws_grid.column_dimensions[chr(64+col)].width = 42
        for r in range(2, 2+len(slots)):
            ws_grid.row_dimensions[r].height = 95
            for c in range(2, 7):
                ws_grid.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")

        # Coverage gaps
        ws_gaps = wb.create_sheet(f"Week{w+1}_CoverageGaps")
        ws_gaps.append(["Date", "Time", "Role", "Issue"])
        for c in ws_gaps[1]:
            c.font = Font(bold=True)
        for d, t, role, issue in gaps:
            ws_gaps.append([d.isoformat(), "" if t is None else t.strftime("%H:%M"), role, issue])

        # Totals
        ws_tot = wb.create_sheet(f"Week{w+1}_Totals")
        rows = []
        for d in dates:
            for t in slots:
                for role, nm in assign.get((d, t), []):
                    rows.append([d, nm, role, 0.5])
                for nm in breaks.get((d, t), set()):
                    rows.append([d, nm, "Break", 0.5])

        df = pd.DataFrame(rows, columns=["Date", "Name", "Task", "Hours"])
        if df.empty:
            df = pd.DataFrame(columns=["Date","Name","Task","Hours"])

        pivot_w = (df.groupby(["Name", "Task"])["Hours"].sum()
                    .reset_index()
                    .pivot(index="Name", columns="Task", values="Hours")
                    .fillna(0.0))
        pivot_w["WeeklyTotal"] = pivot_w.sum(axis=1)
        pivot_w = pivot_w.reset_index()

        pivot_d = df.groupby(["Date", "Name", "Task"])["Hours"].sum().reset_index()

        ws_tot.append(["Weekly totals per staff by task (hours)"])
        ws_tot["A1"].font = Font(bold=True, size=12)
        ws_tot.append([])
        for r in dataframe_to_rows(pivot_w, index=False, header=True):
            ws_tot.append(r)

        header_row = 3
        for cell in ws_tot[header_row]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        headers = [c.value for c in ws_tot[header_row]]
        for j, h in enumerate(headers, start=1):
            if not h or h == "Name":
                continue
            fill = fill_for(h)
            for irow in range(header_row, header_row + len(pivot_w) + 1):
                ws_tot.cell(irow, j).fill = fill

        ws_tot.append([])
        ws_tot.append(["Daily totals per staff by task (hours)"])
        ws_tot[f"A{ws_tot.max_row}"].font = Font(bold=True, size=12)
        ws_tot.append([])
        for r in dataframe_to_rows(pivot_d, index=False, header=True):
            ws_tot.append(r)

        # Timelines with holiday/sick/bank holiday
        ws_tl = wb.create_sheet(f"Week{w+1}_Timelines_All")
        ws_tl.append(["Date", "Time"] + staff_names)
        for c in ws_tl[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_tl.freeze_panes = "C2"

        # assignment lookup
        a_map = {}
        for d in dates:
            for t in slots:
                for role, nm in assign.get((d, t), []):
                    prev = a_map.get((d, t, nm))
                    if prev is None or prev == "Unassigned":
                        a_map[(d, t, nm)] = role

        for d in dates:
            for t in slots:
                row = [d.strftime("%a %d-%b"), t.strftime("%H:%M")]
                for nm in staff_names:
                    hk = holiday_kind(nm, d, hols)
                    if hk:
                        row.append(hk)
                    elif not is_working(hmap, week_start, d, t, nm):
                        row.append("Not working")
                    elif nm in breaks.get((d, t), set()):
                        row.append("Break")
                    else:
                        row.append(a_map.get((d, t, nm), "Unassigned"))
                ws_tl.append(row)

        for r in range(2, ws_tl.max_row + 1):
            for c in range(3, ws_tl.max_column + 1):
                val = ws_tl.cell(r, c).value
                ws_tl.cell(r, c).fill = fill_for(str(val))
                ws_tl.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")

        # By site tabs
        for site in ["SLGP", "JEN", "BGS"]:
            site_staff = list(staff_df.loc[staff_df["HomeSite"].astype(str).str.upper() == site, "Name"].astype(str))
            if not site_staff:
                continue
            ws_site = wb.create_sheet(f"Week{w+1}_{site}_Timelines")
            ws_site.append(["Date", "Time"] + site_staff)
            for c in ws_site[1]:
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws_site.freeze_panes = "C2"

            for d in dates:
                for t in slots:
                    row = [d.strftime("%a %d-%b"), t.strftime("%H:%M")]
                    for nm in site_staff:
                        hk = holiday_kind(nm, d, hols)
                        if hk:
                            row.append(hk)
                        elif not is_working(hmap, week_start, d, t, nm):
                            row.append("Not working")
                        elif nm in breaks.get((d, t), set()):
                            row.append("Break")
                        else:
                            row.append(a_map.get((d, t, nm), "Unassigned"))
                    ws_site.append(row)

            for r in range(2, ws_site.max_row + 1):
                for c in range(3, ws_site.max_column + 1):
                    val = ws_site.cell(r, c).value
                    ws_site.cell(r, c).fill = fill_for(str(val))
                    ws_site.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")

    return wb

# =========================================================
# Streamlit UI
# =========================================================
st.set_page_config(page_title="Rota Generator v7", layout="wide")
require_password()

st.title("Rota Generator (coverage rules restored)")

uploaded = st.file_uploader("Upload rota template (.xlsx)", type=["xlsx"])

c1, c2 = st.columns(2)
with c1:
    start_date = st.date_input("Week commencing (Monday)", value=date.today())
with c2:
    mode = st.selectbox("Run period", ["1 week", "4 weeks (month)", "Custom # weeks"])

if mode == "Custom # weeks":
    weeks = int(st.number_input("Number of weeks", min_value=1, max_value=12, value=2, step=1))
elif mode == "4 weeks (month)":
    weeks = 4
else:
    weeks = 1

start_monday = start_date - timedelta(days=start_date.weekday())

st.caption(
    "Rules: mandatory coverage as specified; min stint 2h; max stint 3h (except Email); "
    "breaks between 12:00–14:00 near midpoint; "
    "cross-site only if no home-site staff available (never for Front Desk; Bookings SLGP-only); "
    "timeline shows Holiday/Sick/Bank Holiday based on Holidays Notes."
)

if uploaded:
    try:
        staff_df, hours_df, hols = read_template(uploaded.getvalue())
        st.success("Template loaded.")
        if st.button("Generate rota and download Excel", type="primary"):
            wb = build_workbook(staff_df, hours_df, hols, start_monday, weeks)
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            out_name = f"rota_{start_monday.isoformat()}_{weeks}w.xlsx"
            st.download_button(
                "📊 Download Excel rota",
                data=bio.getvalue(),
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    except Exception as e:
        st.error("Could not process the template.")
        st.exception(e)
else:
    st.info("Upload your completed template to continue.")
